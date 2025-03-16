import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import glob
import time
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Find the most recent data file
def find_latest_data_file():
    logging.info("Searching for the most recent data file...")
    files = glob.glob("shares_data_till_*.xlsx")
    
    if not files:
        logging.warning("No files found matching the pattern 'shares_data_till_*.xlsx'")
        return None
        
    logging.info(f"Found {len(files)} files matching the pattern")
    
    latest_file = None
    latest_date = None
    
    for file in files:
        try:
            # Extract date part from filename
            date_str = file.replace("shares_data_till_", "").replace(".xlsx", "")
            date_parts = date_str.split("_")
            logging.debug(f"Processing file: {file}, date parts: {date_parts}")
            
            if len(date_parts) == 3:  # Assuming format is YYYY_MM_DD
                file_date = f"{date_parts[0]}-{date_parts[1]}-{date_parts[2]}"
                logging.debug(f"Extracted date: {file_date}")
                
                if latest_date is None or file_date > latest_date:
                    latest_date = file_date
                    latest_file = file
                    logging.debug(f"New latest file: {latest_file} with date {latest_date}")
        except Exception as e:
            logging.warning(f"Error processing filename {file}: {str(e)}")
            continue
    
    if latest_file:
        logging.info(f"Selected latest file: {latest_file} (date: {latest_date})")
    else:
        logging.warning("Could not determine the latest file")
        
    return latest_file

def convert_to_numeric(x):
    """Convert string values with commas to float"""
    if isinstance(x, str):
        return float(x.replace(',', ''))
    return x

def calculate_bollinger_bands(df, period=20, num_std=2):
    """Calculate Bollinger Bands with given period and standard deviation"""
    logging.debug(f"Calculating Bollinger Bands with period {period} and {num_std} std")
    
    # Ensure price is numeric
    numeric_price = df['Ltp_numeric']
    
    # Calculate SMA
    sma = numeric_price.rolling(window=period).mean()
    
    # Calculate standard deviation
    std = numeric_price.rolling(window=period).std()
    
    # Calculate upper and lower Bollinger Bands
    upper_band = sma + (std * num_std)
    lower_band = sma - (std * num_std)
    
    # Set NaN for initial periods where we don't have enough data
    sma.iloc[:period-1] = np.nan
    upper_band.iloc[:period-1] = np.nan
    lower_band.iloc[:period-1] = np.nan
    
    return sma, upper_band, lower_band

def check_bollinger_band_crosses(df, days_to_check=3):
    """Check if OHLC values cross Bollinger Bands in ALL of the last N days"""
    logging.debug(f"Checking for Bollinger Band crossovers in ALL of the last {days_to_check} days")
    
    # Make sure we don't try to check more days than we have data for
    days_to_check = min(days_to_check, len(df))
    
    # Get the data for the last N days
    last_days = df.iloc[-days_to_check:]
    
    # Track crossings for each day
    days_with_upper_cross = 0
    days_with_lower_cross = 0
    band_crossed = "None"
    
    # Check each of the last N days
    for _, day in last_days.iterrows():
        # Ensure all values are numeric
        upper_band = convert_to_numeric(day['Upper_Band'])
        lower_band = convert_to_numeric(day['Lower_Band'])
        open_val = convert_to_numeric(day['Open'])
        high_val = convert_to_numeric(day['High'])
        low_val = convert_to_numeric(day['Low'])
        ltp_val = convert_to_numeric(day['Ltp'])
        
        # Check upper band crosses for this day
        if (open_val > upper_band or high_val > upper_band or 
            low_val > upper_band or ltp_val > upper_band):
            days_with_upper_cross += 1
            
        # Check lower band crosses for this day
        if (open_val < lower_band or high_val < lower_band or 
            low_val < lower_band or ltp_val < lower_band):
            days_with_lower_cross += 1
    
    # Determine which band was crossed in ALL days (if any)
    if days_with_upper_cross == days_to_check:
        band_crossed = "Upper"
        crossed_all_days = True
    elif days_with_lower_cross == days_to_check:
        band_crossed = "Lower"
        crossed_all_days = True
    else:
        crossed_all_days = False
    
    logging.debug(f"Bollinger Band crossover results - Crossed all {days_to_check} days: {crossed_all_days}, " +
                 f"Band crossed: {band_crossed}, Upper crosses: {days_with_upper_cross}, Lower crosses: {days_with_lower_cross}")
    
    return crossed_all_days, band_crossed

def create_bollinger_chart(df, sheet_name, output_dir):
    """Create a chart showing Bollinger Bands with candlestick pattern and save it as an image"""
    logging.info(f"Creating Bollinger Band candlestick chart for {sheet_name}")
    plt.figure(figsize=(10, 6))
    
    # Use a more attractive color scheme and styling
    plt.style.use('ggplot')
    
    # Get date range for the chart
    dates = df['Date']
    
    # Create the candlestick chart
    from mplfinance.original_flavor import candlestick_ohlc
    import matplotlib.dates as mdates
    
    # Format the data for candlestick chart
    ohlc_data = []
    for i, row in df.iterrows():
        # Convert date to matplotlib date number format
        date_num = mdates.date2num(row['Date'])
        # Append (date, open, high, low, close) data
        ohlc_data.append([date_num, 
                           float(row['Open']), 
                           float(row['High']), 
                           float(row['Low']), 
                           float(row['Ltp_numeric'])])  # Using Ltp as close
    
    # Plot the candlestick chart
    ax = plt.gca()
    candlestick_ohlc(ax, ohlc_data, width=0.6, colorup='#2ca02c', colordown='#d62728')
    
    # Format x-axis as dates
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.xticks(rotation=45)
    
    # Plot Bollinger Bands with better colors
    plt.plot(df['Date'], df['SMA_20'], label='20-day SMA', color='#1f77b4', linewidth=2)
    plt.plot(df['Date'], df['Upper_Band'], label='Upper Band', color='#2ca02c', linewidth=1.5)
    plt.plot(df['Date'], df['Lower_Band'], label='Lower Band', color='#d62728', linewidth=1.5)
    
    # Improve chart appearance
    plt.title(f'{sheet_name} - Bollinger Bands Analysis', fontsize=14, fontweight='bold')
    plt.xlabel('Date', fontsize=12)
    plt.ylabel('Price', fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.legend(loc='best', frameon=True, fancybox=True, shadow=True)
    
    # Fill the area between bands with light color
    plt.fill_between(df['Date'], df['Upper_Band'], df['Lower_Band'], color='#b3e0ff', alpha=0.3)
    
    # Add some padding around the plot
    plt.tight_layout()
    
    # Create directory if it doesn't exist
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
    
    # Save chart as image with higher DPI for better quality
    chart_path = os.path.join(charts_dir, f"{sheet_name}_bollinger_chart.png")
    plt.savefig(chart_path, dpi=120)
    plt.close()
    logging.info(f"Chart saved to {chart_path}")
    
    return chart_path

def apply_header_style(cell):
    """Apply header style to a cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_button_style(cell):
    """Apply button style to a cell"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_yes_style(cell):
    """Apply yes (green) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_no_style(cell):
    """Apply no (red) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_cell_style(cell):
    """Apply basic cell style"""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_hyperlink_style(cell):
    """Apply hyperlink style to a cell"""
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def generate_analysis_report(input_file):
    """Generate analysis report for all sheets in the input file"""
    start_time = time.time()
    logging.info(f"Starting Bollinger Bands analysis of file: {input_file}")
    
    # Create output directory structure with today's date
    base_output_dir = os.path.join(os.path.dirname(os.getcwd()), "4. analysis_result")
    bollinger_dir = os.path.join(base_output_dir, "bollinger")
    
    # Create date-based subfolder
    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.join(bollinger_dir, today)
    
    # Create necessary directories
    if not os.path.exists(output_dir):
        logging.info(f"Creating output directory: {output_dir}")
        os.makedirs(output_dir)
    else:
        logging.info(f"Using existing output directory: {output_dir}")
    
    # Read the Excel file all at once
    logging.info("Reading all sheets from Excel file")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    logging.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}") 
    
    # Create a workbook for the analysis results
    results_file = os.path.join(output_dir, "bollinger_bands_analysis_results.xlsx")
    logging.info(f"Creating results workbook: {results_file}")
    results_wb = Workbook()
    
    # Create the summary sheet
    summary_ws = results_wb.active
    summary_ws.title = "Bollinger Summary"
    
    # Add title and date to the summary sheet
    title_cell = summary_ws.cell(row=1, column=1, value="Bollinger Bands Analysis Summary")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    summary_ws.merge_cells('A1:E1')
    
    # Add date information
    date_cell = summary_ws.cell(row=2, column=1, value=f"Generated on: {today}")
    date_cell.font = Font(italic=True)
    
    # Add headers to the summary worksheet
    logging.info("Adding headers to summary worksheet")
    headers = ["Stock", "Latest Price", "Crossed Band in ALL Last 3 Days", "Band Type", "Chart Link"]
    for col, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Set column widths
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 30
    summary_ws.column_dimensions['D'].width = 15
    summary_ws.column_dimensions['E'].width = 15
    
    # Process each sheet
    row = 5  # Start data from row 5
    successful_sheets = 0
    failed_sheets = 0
    
    # Prepare data for detailed sheet
    detailed_data = []
    
    # Start processing sheets
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
        
    for sheet_idx, sheet_name in enumerate(sheet_names):
        try:
            logging.info(f"Processing sheet {sheet_idx+1}/{len(sheet_names)}: {sheet_name}")
            
            # Read data from the sheet
            logging.info(f"Reading data from sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logging.info(f"Data shape: {df.shape} rows × {df.shape[1]} columns")
            
            # Convert Date column to datetime if needed
            logging.info("Converting Date column to datetime format")
            if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
                df['Date'] = pd.to_datetime(df['Date'])
            
            # Convert Ltp to numeric by removing commas
            logging.info("Converting Ltp values to numeric format")
            df['Ltp_numeric'] = df['Ltp'].apply(convert_to_numeric)
            
            # Make sure OHLC columns are numeric
            ohlc_columns = ['Open', 'High', 'Low', 'Close']
            for col in ohlc_columns:
                if col in df.columns:
                    df[col] = df[col].apply(convert_to_numeric)
            
            # Calculate Bollinger Bands
            logging.info("Calculating Bollinger Bands")
            df['SMA_20'], df['Upper_Band'], df['Lower_Band'] = calculate_bollinger_bands(df, period=20, num_std=2)
            
            # Check for Bollinger Band crossovers in ALL last 3 days
            logging.info("Checking for Bollinger Band crossovers in ALL last 3 days")
            crosses_band, band_crossed = check_bollinger_band_crosses(df, days_to_check=3)
            logging.info(f"Bollinger Band crossover results - Crossed ALL 3 days: {crosses_band}, Band crossed: {band_crossed}")
            
            # Create chart for this stock
            chart_path = create_bollinger_chart(df, sheet_name, output_dir)
            
            # Get latest day data for summary
            latest_day = df.iloc[-1]
            
            # Get latest price
            latest_price = float(latest_day['Ltp']) if isinstance(latest_day['Ltp'], (int, float)) else float(str(latest_day['Ltp']).replace(',', ''))
            
            # Add results to the summary worksheet
            logging.info("Adding results to summary worksheet")
            cell_stock = summary_ws.cell(row=row, column=1, value=sheet_name)
            apply_cell_style(cell_stock)
            
            # Add latest price
            cell_price = summary_ws.cell(row=row, column=2, value=f"{latest_price:.2f}")
            apply_cell_style(cell_price)
            
            # Set Yes/No values with styles for band crossing
            cell_crosses = summary_ws.cell(row=row, column=3, value="Yes" if crosses_band else "No")
            if crosses_band:
                apply_yes_style(cell_crosses)
            else:
                apply_no_style(cell_crosses)
            
            # Add which band was crossed
            cell_band_type = summary_ws.cell(row=row, column=4, value=band_crossed if crosses_band else "None")
            apply_cell_style(cell_band_type)
            if band_crossed == "Upper":
                cell_band_type.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif band_crossed == "Lower":
                cell_band_type.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Light red
            
            # Add chart link placeholder
            cell_chart = summary_ws.cell(row=row, column=5, value="View Chart")
            apply_cell_style(cell_chart)
            
            # Store data for detailed sheet
            latest_date = latest_day['Date']
            if isinstance(latest_date, pd.Timestamp):
                latest_date = latest_date.strftime('%Y-%m-%d')
                
            detailed_data.append({
                'stock': sheet_name,
                'date': latest_date,
                'price': round(latest_price, 2),
                'upper_band': round(latest_day['Upper_Band'], 2),
                'lower_band': round(latest_day['Lower_Band'], 2),
                'band_crossed': band_crossed
            })
            
            row += 1
            successful_sheets += 1
            logging.info(f"Successfully processed sheet: {sheet_name}")
            
        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {str(e)}")
            failed_sheets += 1
            continue
    
    # Add alternating row colors for readability
    for row_idx in range(5, row):
        if row_idx % 2 == 0:  # Even rows
            for col in range(1, 6):
                cell = summary_ws.cell(row=row_idx, column=col)
                if not cell.fill or cell.fill.fill_type != "solid":
                    cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    
    # Add summary information
    summary_cell = summary_ws.cell(row=3, column=1, value=f"Total Stocks Analyzed: {row-5}")
    summary_cell.font = Font(bold=True)
    summary_ws.merge_cells('A3:E3')
    
    # Freeze panes for better navigation
    summary_ws.freeze_panes = summary_ws['A5']
    
    # Save the results workbook
    logging.info(f"Saving results workbook to {results_file}")
    results_wb.save(results_file)
    
    # Add hyperlinks to charts and navigation buttons
    logging.info("Adding hyperlinks to charts and navigation buttons")
    add_chart_hyperlinks_and_navigation(results_file, output_dir)
    
    # Create a detailed sheet with all Bollinger Band data
    logging.info("Creating detailed Bollinger Band data sheet")
    create_detailed_bollinger_sheet(results_file, detailed_data)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Analysis complete in {elapsed_time:.2f} seconds")
    logging.info(f"Successfully processed {successful_sheets} sheets, failed to process {failed_sheets} sheets")
    logging.info(f"Results saved to {results_file}")

def create_detailed_bollinger_sheet(results_file, detailed_data):
    """Create a detailed sheet with Bollinger Band data for each stock"""
    logging.info("Processing detailed Bollinger Band data")
    
    try:
        # Load the workbook
        wb = load_workbook(results_file)
        
        # Create a new sheet for detailed data
        detailed_sheet = wb.create_sheet(title="Detailed Data")
        
        # Add title
        title_cell = detailed_sheet.cell(row=1, column=1, value="Detailed Bollinger Bands Data")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        detailed_sheet.merge_cells('A1:F1')
        
        # Add headers
        headers = ["Stock", "Latest Date", "Latest Price", "Upper Band", "Lower Band", "Band Crossed"]
        for col, header in enumerate(headers, 1):
            cell = detailed_sheet.cell(row=3, column=col, value=header)
            apply_header_style(cell)
        
        # Set column widths
        detailed_sheet.column_dimensions['A'].width = 25
        detailed_sheet.column_dimensions['B'].width = 15
        detailed_sheet.column_dimensions['C'].width = 15
        detailed_sheet.column_dimensions['D'].width = 15
        detailed_sheet.column_dimensions['E'].width = 15
        detailed_sheet.column_dimensions['F'].width = 15
        
        # Add data to the detailed sheet
        row = 4
        for entry in detailed_data:
            detailed_sheet.cell(row=row, column=1, value=entry['stock'])
            detailed_sheet.cell(row=row, column=2, value=entry['date'])
            detailed_sheet.cell(row=row, column=3, value=entry['price'])
            detailed_sheet.cell(row=row, column=4, value=entry['upper_band'])
            detailed_sheet.cell(row=row, column=5, value=entry['lower_band'])
            
            # Band crossed info
            cell_band = detailed_sheet.cell(row=row, column=6, value=entry['band_crossed'])
            if entry['band_crossed'] == "Upper":
                cell_band.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif entry['band_crossed'] == "Lower":
                cell_band.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Light red
            
            # Apply styles to all cells in the row
            for col in range(1, 7):
                apply_cell_style(detailed_sheet.cell(row=row, column=col))
            
            row += 1
        
        # Apply alternating row colors
        for row_idx in range(4, row):
            if row_idx % 2 == 0:  # Even rows
                for col in range(1, 7):
                    cell = detailed_sheet.cell(row=row_idx, column=col)
                    if not cell.fill or cell.fill.fill_type != "solid":
                        cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        
        # Add navigation button back to main summary
        button_cell = detailed_sheet.cell(row=2, column=6, value="Back to Summary")
        apply_button_style(button_cell)
        button_cell.hyperlink = f"#'Bollinger Summary'!A1"
        
        # Freeze panes
        detailed_sheet.freeze_panes = detailed_sheet['A4']
        
        # Save the workbook
        logging.info("Saving workbook with detailed data sheet")
        wb.save(results_file)
        logging.info("Detailed Bollinger Band data sheet created successfully")
        
    except Exception as e:
        logging.error(f"Error creating detailed data sheet: {str(e)}")

def add_chart_hyperlinks_and_navigation(results_file, output_dir):
    """Add hyperlinks to chart images and navigation buttons in the Excel file"""
    logging.info("Adding chart hyperlinks and navigation buttons to the Excel file")
    try:
        wb = load_workbook(results_file)
        main_sheet = wb["Bollinger Summary"]
        
        # Get the maximum row
        max_row = main_sheet.max_row
        logging.info(f"Processing {max_row-4} stock entries")  # Subtract header rows
        
        # Create a sheet for each chart
        for row in range(5, max_row + 1):  # Start from row 5 (data starts here)
            stock_name = main_sheet[f'A{row}'].value
            if not stock_name:
                continue
                
            chart_path = os.path.join(output_dir, "charts", f"{stock_name}_bollinger_chart.png")
            
            if os.path.exists(chart_path):
                logging.debug(f"Creating chart sheet for {stock_name}")
                # Create a new sheet for this stock's chart
                chart_sheet_name = f"{stock_name}_Chart"
                # Use truncated name if too long (Excel has a 31 character limit for sheet names)
                if len(chart_sheet_name) > 31:
                    chart_sheet_name = chart_sheet_name[:28] + "..."
                    
                chart_sheet = wb.create_sheet(title=chart_sheet_name)
                
                # Add a title to the chart sheet
                chart_sheet['A1'] = f"{stock_name} - Bollinger Bands Analysis"
                chart_sheet['A1'].font = Font(bold=True, size=14)
                chart_sheet['A1'].alignment = Alignment(horizontal="center")
                
                # Merge cells for the title
                chart_sheet.merge_cells('A1:F1')
                
                # Add the chart image to the sheet
                img = Image(chart_path)
                img.width = 800
                img.height = 500
                chart_sheet.add_image(img, 'A3')
                
                # Add "Go to Summary" button at position O18 (right side of image at mid-height)
                button_cell = chart_sheet['O18'] = "Go to Summary"
                apply_button_style(chart_sheet['O18'])
                chart_sheet['O18'].hyperlink = f"#'Bollinger Summary'!A1"
                
                # Make sure column O is wide enough for button text
                chart_sheet.column_dimensions['O'].width = 20
                
                # Create a hyperlink from the results sheet to the chart sheet
                link_cell = main_sheet[f'E{row}']
                link_cell.value = "View Chart"
                link_cell.hyperlink = f"#{chart_sheet_name}!A1"
                apply_hyperlink_style(link_cell)
            else:
                logging.warning(f"Chart not found for {stock_name}: {chart_path}")
        
        logging.info(f"Saving workbook with hyperlinks and navigation to {results_file}")
        wb.save(results_file)
        logging.info("Chart hyperlinks and navigation buttons added successfully")
    except Exception as e:
        logging.error(f"Error adding chart hyperlinks and navigation: {str(e)}")

if __name__ == "__main__":
    logging.info("===== BOLLINGER BANDS ANALYSIS SCRIPT STARTED =====")
    
    # Find the latest data file
    latest_file = find_latest_data_file()
    
    if latest_file:
        logging.info(f"Processing file: {latest_file}")
        generate_analysis_report(latest_file)
    else:
        logging.error("No data files found matching the pattern 'shares_data_till_*.xlsx'")
    
    logging.info("===== BOLLINGER BANDS ANALYSIS SCRIPT COMPLETED =====")