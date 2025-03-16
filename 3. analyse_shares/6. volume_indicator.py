import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import glob
import time
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Find the most recent data file
def find_latest_data_file():
    logging.info("Searching for the most recent data file...")
    files = glob.glob("shares_data_till_*.xlsx")
    
    if not files:
        logging.warning("No files found matching the pattern 'shares_data_till_*.xlsx'")
        return None
        
    logging.info(f"Found {len(files)} files matching the pattern")
    
    latest_file = None
    latest_date = None
    
    for file in files:
        try:
            # Extract date part from filename
            date_str = file.replace("shares_data_till_", "").replace(".xlsx", "")
            date_parts = date_str.split("_")
            logging.debug(f"Processing file: {file}, date parts: {date_parts}")
            
            if len(date_parts) == 3:  # Assuming format is YYYY_MM_DD
                file_date = f"{date_parts[0]}-{date_parts[1]}-{date_parts[2]}"
                logging.debug(f"Extracted date: {file_date}")
                
                if latest_date is None or file_date > latest_date:
                    latest_date = file_date
                    latest_file = file
                    logging.debug(f"New latest file: {latest_file} with date {latest_date}")
        except Exception as e:
            logging.warning(f"Error processing filename {file}: {str(e)}")
            continue
    
    if latest_file:
        logging.info(f"Selected latest file: {latest_file} (date: {latest_date})")
    else:
        logging.warning("Could not determine the latest file")
        
    return latest_file

def convert_to_numeric(x):
    """Convert string values with commas to float"""
    if isinstance(x, str):
        return float(x.replace(',', ''))
    return x

def analyze_volume(df, recent_days=3, previous_days=5):
    """
    Analyze if recent volume is higher than the historical average
    Compares average volume of latest 3 days with average of previous 5 days
    
    Parameters:
    - df: DataFrame containing stock data
    - recent_days: Number of most recent days to average (default: 3)
    - previous_days: Number of days before that to average (default: 5)
    
    Returns:
    - is_volume_increasing: Boolean indicating if volume is increasing
    - pct_change: Percentage change in volume
    """
    logging.debug("Analyzing volume data")
    
    # Make sure Qty is numeric
    df['Qty_numeric'] = df['Qty'].apply(convert_to_numeric)
    
    # Check if we have enough data
    if len(df) < recent_days + previous_days:
        logging.warning(f"Not enough data points for volume analysis. Need at least {recent_days + previous_days} points.")
        return False, 0.0
    
    # Calculate average volume for the most recent 3 days
    recent_avg_volume = df['Qty_numeric'].tail(recent_days).mean()
    
    # Calculate average volume for the previous 5 days before the recent 3 days
    previous_avg_volume = df['Qty_numeric'].iloc[-(recent_days+previous_days):-recent_days].mean()
    
    # Check if recent volume is higher
    is_volume_increasing = recent_avg_volume > previous_avg_volume
    
    # Calculate percentage increase/decrease
    pct_change = ((recent_avg_volume - previous_avg_volume) / previous_avg_volume) * 100
    
    logging.debug(f"Recent {recent_days}-day avg volume: {recent_avg_volume:.2f}")
    logging.debug(f"Previous {previous_days}-day avg volume: {previous_avg_volume:.2f}")
    logging.debug(f"Volume is {'increasing' if is_volume_increasing else 'decreasing'} by {abs(pct_change):.2f}%")
    
    return is_volume_increasing, pct_change

def create_volume_chart(df, sheet_name, output_dir):
    """Create a chart showing volume trends and save it as an image"""
    logging.info(f"Creating volume chart for {sheet_name}")
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Use a more attractive color scheme and styling
    plt.style.use('ggplot')
    
    # Make sure necessary columns are numeric
    df['Qty_numeric'] = df['Qty'].apply(convert_to_numeric)
    
    # Create a serial x-axis - use index positions instead of dates
    # This ensures no gaps in the chart
    x_positions = np.arange(len(df))
    
    # Set date labels for x-axis ticks (we'll use fewer ticks for readability)
    date_labels = df['Date']
    
    # Ensure we have Open and Close columns
    if 'Open' not in df.columns or 'Close' not in df.columns:
        logging.warning(f"Open or Close columns missing in {sheet_name}. Using default volume bar colors.")
        # Plot volume data with default color
        ax.bar(x_positions, df['Qty_numeric'], alpha=0.6, color='#1f77b4', label='Volume')
    else:
        # Convert Open and Close to numeric if they're not already
        df['Open_numeric'] = df['Open'].apply(convert_to_numeric)
        df['Close_numeric'] = df['Close'].apply(convert_to_numeric)
        
        # Create color array based on Open vs Close comparison
        colors = ['#ff5050' if open_price > close_price else '#33cc33' 
                 for open_price, close_price in zip(df['Open_numeric'], df['Close_numeric'])]
        
        # Plot volume bars with conditional colors
        ax.bar(x_positions, df['Qty_numeric'], alpha=0.7, color=colors, label='Volume')
        
        # Create custom legend patches for red and green bars
        from matplotlib.patches import Patch
        red_patch = Patch(color='#ff5050', alpha=0.7, label='Volume (Price Decrease)')
        green_patch = Patch(color='#33cc33', alpha=0.7, label='Volume (Price Increase)')
        legend_elements = [red_patch, green_patch]
    
    # Add moving average lines for volume - now using serial positions
    volume_ma10 = df['Qty_numeric'].rolling(window=10).mean()
    volume_ma20 = df['Qty_numeric'].rolling(window=20).mean()
    
    ax.plot(x_positions, volume_ma10, color='#ff7f0e', linewidth=2, label='10-day MA')
    ax.plot(x_positions, volume_ma20, color='#d62728', linewidth=2, label='20-day MA')
    
    # Improve chart appearance
    ax.set_title(f'{sheet_name} - Volume Analysis', fontsize=14, fontweight='bold')
    ax.set_xlabel('Trading Sessions', fontsize=12)
    ax.set_ylabel('Volume', fontsize=12)
    ax.grid(True, alpha=0.3)
    
    # Configure x-axis ticks to show dates at reasonable intervals
    # If we have many data points, we'll show fewer tick labels
    num_points = len(df)
    if num_points > 50:
        # For large datasets, show approximately 10 dates
        step = num_points // 10
        if step < 1:
            step = 1
        tick_positions = x_positions[::step]
        tick_labels = [date.strftime('%Y-%m-%d') if isinstance(date, pd.Timestamp) 
                       else str(date) for date in date_labels.iloc[::step]]
    else:
        # For smaller datasets, show more frequent date labels
        tick_positions = x_positions
        tick_labels = [date.strftime('%Y-%m-%d') if isinstance(date, pd.Timestamp) 
                      else str(date) for date in date_labels]
    
    ax.set_xticks(tick_positions)
    ax.set_xticklabels(tick_labels, rotation=45, ha='right')
    
    # Add legend with proper elements
    if 'Open' in df.columns and 'Close' in df.columns:
        ax.legend(handles=legend_elements + [
            plt.Line2D([0], [0], color='#ff7f0e', linewidth=2, label='10-day MA'),
            plt.Line2D([0], [0], color='#d62728', linewidth=2, label='20-day MA')
        ], loc='best', frameon=True, fancybox=True, shadow=True)
    else:
        ax.legend(loc='best', frameon=True, fancybox=True, shadow=True)
    
    # Format y-axis to show volume in millions or thousands
    max_volume = df['Qty_numeric'].max()
    if max_volume > 1000000:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x/1000000:.1f}M'))
        ax.set_ylabel('Volume (Millions)', fontsize=12)
    elif max_volume > 1000:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x/1000:.1f}K'))
        ax.set_ylabel('Volume (Thousands)', fontsize=12)
    
    # Add some padding around the plot
    plt.tight_layout()
    
    # Create directory if it doesn't exist
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
    
    # Save chart as image with higher DPI for better quality
    chart_path = os.path.join(charts_dir, f"{sheet_name}_volume_chart.png")
    plt.savefig(chart_path, dpi=120)
    plt.close()
    logging.info(f"Volume chart saved to {chart_path}")
    
    return chart_path

def apply_header_style(cell):
    """Apply header style to a cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_button_style(cell):
    """Apply button style to a cell"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_yes_style(cell):
    """Apply yes (green) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_no_style(cell):
    """Apply no (red) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_cell_style(cell):
    """Apply basic cell style"""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_hyperlink_style(cell):
    """Apply hyperlink style to a cell"""
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def generate_volume_analysis_report(input_file):
    """Generate volume analysis report for all sheets in the input file"""
    start_time = time.time()
    logging.info(f"Starting volume analysis of file: {input_file}")
    
    # Create output directory structure with today's date
    base_output_dir = os.path.join(os.path.dirname(os.getcwd()), "4. analysis_result")
    volume_dir = os.path.join(base_output_dir, "volume")
    
    # Create date-based subfolder
    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.join(volume_dir, today)
    
    # Create necessary directories
    if not os.path.exists(output_dir):
        logging.info(f"Creating output directory: {output_dir}")
        os.makedirs(output_dir)
    else:
        logging.info(f"Using existing output directory: {output_dir}")
    
    # Read the Excel file
    logging.info("Reading Excel file and getting sheet names")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    logging.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
    
    # Create a workbook for the analysis results
    results_file = os.path.join(output_dir, "volume_analysis_results.xlsx")
    logging.info(f"Creating results workbook: {results_file}")
    results_wb = Workbook()
    results_ws = results_wb.active
    results_ws.title = "Volume Analysis Results"
    
    # Add title and date to the main sheet
    title_cell = results_ws.cell(row=1, column=1, value="Stock Volume Analysis")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    results_ws.merge_cells('A1:D1')
    
    # Add date information
    date_cell = results_ws.cell(row=2, column=1, value=f"Generated on: {today}")
    date_cell.font = Font(italic=True)
    
    # Add headers to the results worksheet
    logging.info("Adding headers to results worksheet")
    headers = ["Stock", "Increasing Volume", "Volume Change %", "Chart Link"]
    for col, header in enumerate(headers, 1):
        cell = results_ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Set column widths
    results_ws.column_dimensions['A'].width = 25
    results_ws.column_dimensions['B'].width = 20
    results_ws.column_dimensions['C'].width = 20
    results_ws.column_dimensions['D'].width = 15
    
    # Process each sheet
    row = 5  # Start data from row 5
    successful_sheets = 0
    failed_sheets = 0
    
    for sheet_idx, sheet_name in enumerate(sheet_names):
        try:
            logging.info(f"Processing sheet {sheet_idx+1}/{len(sheet_names)}: {sheet_name}")
            
            # Read data from the sheet
            logging.info(f"Reading data from sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logging.info(f"Data shape: {df.shape} rows × {df.shape[1]} columns")
            
            # Convert Date column to datetime if needed
            logging.info("Converting Date column to datetime format")
            if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
                df['Date'] = pd.to_datetime(df['Date'])
            
            # Ensure we have the Qty column
            if 'Qty' not in df.columns:
                logging.warning(f"Sheet {sheet_name} does not have a Qty column. Skipping.")
                continue
            
            # Sort by date to ensure chronological order
            df = df.sort_values('Date')
            
            # Analyze volume - check if volume is increasing
            logging.info("Analyzing volume trends")
            is_volume_increasing, volume_change_pct = analyze_volume(df)
            
            # Create chart for this stock
            chart_path = create_volume_chart(df, sheet_name, output_dir)
            
            # Add results to the results worksheet
            logging.info("Adding results to worksheet")
            cell_stock = results_ws.cell(row=row, column=1, value=sheet_name)
            apply_cell_style(cell_stock)
            
            # Set Yes/No values with styles
            cell_vol_inc = results_ws.cell(row=row, column=2, value="Yes" if is_volume_increasing else "No")
            if is_volume_increasing:
                apply_yes_style(cell_vol_inc)
            else:
                apply_no_style(cell_vol_inc)
            
            # Add volume change percentage
            cell_vol_change = results_ws.cell(row=row, column=3, value=f"{volume_change_pct:.2f}%")
            apply_cell_style(cell_vol_change)
            
            # If volume change is positive, use green styling, otherwise red
            if volume_change_pct > 0:
                cell_vol_change.font = Font(color="006400", bold=True)
            else:
                cell_vol_change.font = Font(color="FF0000", bold=True)
            
            # Add chart link placeholder
            cell_chart = results_ws.cell(row=row, column=4, value="View Chart")
            apply_cell_style(cell_chart)
            
            row += 1
            successful_sheets += 1
            logging.info(f"Successfully processed sheet: {sheet_name}")
            
        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {str(e)}")
            failed_sheets += 1
            continue
    
    # Add alternating row colors for readability
    for row_idx in range(5, row):
        if row_idx % 2 == 0:  # Even rows
            for col in range(1, 5):
                cell = results_ws.cell(row=row_idx, column=col)
                if not cell.fill or cell.fill.fill_type != "solid":
                    cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    
    # Add summary information
    summary_cell = results_ws.cell(row=3, column=1, value=f"Total Stocks Analyzed: {row-5}")
    summary_cell.font = Font(bold=True)
    results_ws.merge_cells('A3:D3')
    
    # Freeze panes for better navigation
    results_ws.freeze_panes = results_ws['A5']
    
    # Save the results workbook
    logging.info(f"Saving results workbook to {results_file}")
    results_wb.save(results_file)
    
    # Add hyperlinks to charts and navigation buttons
    logging.info("Adding hyperlinks to charts and navigation buttons")
    add_chart_hyperlinks_and_navigation(results_file, output_dir)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Analysis complete in {elapsed_time:.2f} seconds")
    logging.info(f"Successfully processed {successful_sheets} sheets, failed to process {failed_sheets} sheets")
    logging.info(f"Results saved to {results_file}")

def add_chart_hyperlinks_and_navigation(results_file, output_dir):
    """Add hyperlinks to chart images and navigation buttons in the Excel file"""
    logging.info("Adding chart hyperlinks and navigation buttons to the Excel file")
    wb = load_workbook(results_file)
    main_sheet = wb.active
    
    # Get the maximum row
    max_row = main_sheet.max_row
    logging.info(f"Processing {max_row-4} stock entries")  # Subtract header rows
    
    # Create a sheet for each chart
    for row in range(5, max_row + 1):  # Start from row 5 (data starts here)
        stock_name = main_sheet[f'A{row}'].value
        if not stock_name:
            continue
            
        chart_path = os.path.join(output_dir, "charts", f"{stock_name}_volume_chart.png")
        
        if os.path.exists(chart_path):
            logging.info(f"Creating chart sheet for {stock_name}")
            # Create a new sheet for this stock's chart
            chart_sheet_name = f"{stock_name}_Volume"
            # Use truncated name if too long (Excel has a 31 character limit for sheet names)
            if len(chart_sheet_name) > 31:
                chart_sheet_name = chart_sheet_name[:28] + "..."
                
            chart_sheet = wb.create_sheet(title=chart_sheet_name)
            
            # Add a title to the chart sheet
            chart_sheet['A1'] = f"{stock_name} - Volume Trend Analysis"
            chart_sheet['A1'].font = Font(bold=True, size=14)
            chart_sheet['A1'].alignment = Alignment(horizontal="center")
            
            # Merge cells for the title
            chart_sheet.merge_cells('A1:F1')
            
            # Add the chart image to the sheet
            logging.info(f"Adding chart image to sheet: {chart_sheet_name}")
            img = Image(chart_path)
            img.width = 800
            img.height = 500
            chart_sheet.add_image(img, 'A3')
            
            # Add "Go to Main Page" button at position O18 (right side of image at mid-height)
            button_cell = chart_sheet['O18'] = "Go to Main Page"
            apply_button_style(chart_sheet['O18'])
            chart_sheet['O18'].hyperlink = f"#'Volume Analysis Results'!A1"
            
            # Make sure column O is wide enough for button text
            chart_sheet.column_dimensions['O'].width = 20
            
            # Create a hyperlink from the results sheet to the chart sheet
            logging.info(f"Creating hyperlink for {stock_name}")
            link_cell = main_sheet[f'D{row}']
            link_cell.value = "View Chart"
            link_cell.hyperlink = f"#{chart_sheet_name}!A1"
            apply_hyperlink_style(link_cell)
        else:
            logging.warning(f"Chart not found for {stock_name}: {chart_path}")
    
    logging.info(f"Saving workbook with hyperlinks and navigation to {results_file}")
    wb.save(results_file)
    logging.info("Chart hyperlinks and navigation buttons added successfully")

if __name__ == "__main__":
    logging.info("===== STOCK VOLUME ANALYSIS SCRIPT STARTED =====")
    
    # Find the latest data file
    latest_file = find_latest_data_file()
    
    if latest_file:
        logging.info(f"Processing file: {latest_file}")
        generate_volume_analysis_report(latest_file)
    else:
        logging.error("No data files found matching the pattern 'shares_data_till_*.xlsx'")
    
    logging.info("===== STOCK VOLUME ANALYSIS SCRIPT COMPLETED =====")