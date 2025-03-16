import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import glob
import time
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Find the most recent data file
def find_latest_data_file():
    logging.info("Searching for the most recent data file...")
    files = glob.glob("shares_data_till_*.xlsx")
    
    if not files:
        logging.warning("No files found matching the pattern 'shares_data_till_*.xlsx'")
        return None
        
    logging.info(f"Found {len(files)} files matching the pattern")
    
    latest_file = None
    latest_date = None
    
    for file in files:
        try:
            # Extract date part from filename
            date_str = file.replace("shares_data_till_", "").replace(".xlsx", "")
            logging.debug(f"Processing file: {file}, date string: {date_str}")
            
            # For yyyy_mm_dd format (new format)
            if len(date_str.split("_")) == 3:
                year, month, day = date_str.split("_")
                file_date = f"{year}-{month}-{day}"
                logging.debug(f"Extracted date (new format): {file_date}")
                
                if latest_date is None or file_date > latest_date:
                    latest_date = file_date
                    latest_file = file
                    logging.debug(f"New latest file: {latest_file} with date {latest_date}")
        except Exception as e:
            logging.warning(f"Error processing filename {file}: {str(e)}")
            continue
    
    if latest_file:
        logging.info(f"Selected latest file: {latest_file} (date: {latest_date})")
    else:
        logging.warning("Could not determine the latest file")
        
    return latest_file

def convert_to_numeric(x):
    """Convert string values with commas to float"""
    if isinstance(x, str):
        return float(x.replace(',', ''))
    return x

def calculate_ema(data, period):
    """Calculate Exponential Moving Average with proper initialization"""
    logging.debug(f"Calculating {period}-day EMA")
    # First ensure Ltp is numeric
    numeric_ltp = data['Ltp_numeric']
    ema = numeric_ltp.ewm(span=period, adjust=False).mean()
    
    # Set NaN for initial periods where we don't have enough data
    ema.iloc[:period-1] = np.nan
    
    return ema

def calculate_macd(data):
    """Calculate MACD, Signal Line, and Histogram with proper NaN values"""
    logging.info("Calculating MACD components")
    
    # Calculate 12-day and 26-day EMA
    ema_12 = calculate_ema(data, 12)
    ema_26 = calculate_ema(data, 26)
    
    # Calculate MACD line (12-day EMA - 26-day EMA)
    macd_line = ema_12 - ema_26
    
    # Calculate Signal line (9-day EMA of MACD line)
    signal_line = macd_line.ewm(span=9, adjust=False).mean()
    
    # Properly set NaN values for the Signal Line
    # We need 9-1 additional NaN values after the first MACD value
    first_valid_macd = (~pd.isna(macd_line)).idxmax() if not macd_line.isna().all() else 0
    signal_line.iloc[:first_valid_macd+8] = np.nan
    
    # Calculate MACD Histogram (MACD line - Signal line)
    histogram = macd_line - signal_line
    
    return macd_line, signal_line, histogram

def check_macd_crossover(df, lookback_days=3):
    """Check if MACD line crosses Signal line from below within recent data"""
    logging.debug("Checking for MACD crossovers")
    
    # Get recent data where both MACD and Signal have values
    recent_data = df.dropna(subset=['MACD_Line', 'Signal_Line']).tail(lookback_days)
    
    if len(recent_data) < 2:
        logging.warning("Not enough data points with valid MACD values")
        return False
    
    cross_macd_signal = False
    
    # Check each pair of consecutive rows
    for i in range(1, len(recent_data)):
        prev = recent_data.iloc[i-1]
        curr = recent_data.iloc[i]
        
        # Check MACD crossing Signal from below
        if prev['MACD_Line'] <= prev['Signal_Line'] and curr['MACD_Line'] > curr['Signal_Line']:
            cross_macd_signal = True
            logging.debug(f"MACD crossed Signal from below between {prev.name} and {curr.name}")
            
    logging.debug(f"MACD-Signal crossover result: {cross_macd_signal}")
    return cross_macd_signal

def create_macd_chart(df, sheet_name, output_dir):
    """Create a chart showing MACD, Signal Line and Histogram, and save it as an image"""
    logging.info(f"Creating MACD chart for {sheet_name}")
    
    # Create figure with 2 subplots (price above, MACD below) with correct aspect ratio
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 7), gridspec_kw={'height_ratios': [2, 1]}, sharex=True)
    
    # Use a more attractive color scheme and styling
    plt.style.use('ggplot')
    
    # Plot actual price in top subplot
    ax1.plot(df['Date'], df['Ltp_numeric'], label='Price', color='#404040', alpha=0.6, linewidth=1.5)
    ax1.set_title(f'{sheet_name} - Price Chart', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Price', fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.legend(loc='upper left')
    
    # Plot MACD components in bottom subplot
    ax2.plot(df['Date'], df['MACD_Line'], label='MACD Line', color='#1f77b4', linewidth=1.5)
    ax2.plot(df['Date'], df['Signal_Line'], label='Signal Line', color='#ff7f0e', linewidth=1.5)
    
    # Plot histogram as bar chart
    positive_hist = df['Histogram'].copy()
    negative_hist = df['Histogram'].copy()
    positive_hist[positive_hist <= 0] = np.nan
    negative_hist[negative_hist > 0] = np.nan
    
    ax2.bar(df['Date'], positive_hist, label='Histogram (Positive)', color='#2ca02c', alpha=0.7, width=1)
    ax2.bar(df['Date'], negative_hist, label='Histogram (Negative)', color='#d62728', alpha=0.7, width=1)
    
    # Add zero line
    ax2.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    
    ax2.set_title('MACD (12,26,9)', fontsize=12, fontweight='bold')
    ax2.set_ylabel('MACD', fontsize=10)
    ax2.set_xlabel('Date', fontsize=10)
    ax2.grid(True, alpha=0.3)
    ax2.legend(loc='upper left', fontsize=8)
    
    # Format x-axis dates
    fig.autofmt_xdate()
    
    # Add some padding around the plot
    plt.tight_layout()
    
    # Create directory if it doesn't exist
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
    
    # Save chart as image with higher DPI for better quality and fixed aspect ratio
    chart_path = os.path.join(charts_dir, f"{sheet_name}_macd_chart.png")
    plt.savefig(chart_path, dpi=120, bbox_inches='tight')
    plt.close()
    logging.info(f"MACD chart saved to {chart_path}")
    
    return chart_path

def apply_header_style(cell):
    """Apply header style to a cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_button_style(cell):
    """Apply button style to a cell"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_yes_style(cell):
    """Apply yes (green) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_no_style(cell):
    """Apply no (red) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_cell_style(cell):
    """Apply basic cell style"""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_hyperlink_style(cell):
    """Apply hyperlink style to a cell"""
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def generate_analysis_report(input_file):
    """Generate MACD analysis report for all sheets in the input file"""
    start_time = time.time()
    logging.info(f"Starting MACD analysis of file: {input_file}")
    
    # Create output directory structure with today's date
    base_output_dir = os.path.join(os.path.dirname(os.getcwd()), "4. analysis_result")
    macd_dir = os.path.join(base_output_dir, "macd")
    
    # Create date-based subfolder
    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.join(macd_dir, today)
    
    # Create necessary directories
    if not os.path.exists(output_dir):
        logging.info(f"Creating output directory: {output_dir}")
        os.makedirs(output_dir)
    else:
        logging.info(f"Using existing output directory: {output_dir}")
    
    # Read the Excel file
    logging.info("Reading Excel file and getting sheet names")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    logging.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
    
    # Create a workbook for the analysis results
    results_file = os.path.join(output_dir, "macd_analysis_results.xlsx")
    logging.info(f"Creating results workbook: {results_file}")
    results_wb = Workbook()
    results_ws = results_wb.active
    results_ws.title = "Analysis Results"
    
    # Add title and date to the main sheet
    title_cell = results_ws.cell(row=1, column=1, value="MACD Crossover Analysis")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    results_ws.merge_cells('A1:C1')
    
    # Add date information
    date_cell = results_ws.cell(row=2, column=1, value=f"Generated on: {today}")
    date_cell.font = Font(italic=True)
    
    # Add headers to the results worksheet
    logging.info("Adding headers to results worksheet")
    headers = ["Stock", "MACD crosses Signal Line", "Chart Link"]
    for col, header in enumerate(headers, 1):
        cell = results_ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Set column widths
    results_ws.column_dimensions['A'].width = 25
    results_ws.column_dimensions['B'].width = 25
    results_ws.column_dimensions['C'].width = 15
    
    # Process each sheet
    row = 5  # Start data from row 5
    successful_sheets = 0
    failed_sheets = 0
    
    for sheet_idx, sheet_name in enumerate(sheet_names):
        try:
            logging.info(f"Processing sheet {sheet_idx+1}/{len(sheet_names)}: {sheet_name}")
            
            # Read data from the sheet
            logging.info(f"Reading data from sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logging.info(f"Data shape: {df.shape} rows × {df.shape[1]} columns")
            
            # Convert Date column to datetime if needed
            logging.info("Converting Date column to datetime format")
            if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
                df['Date'] = pd.to_datetime(df['Date'])
            
            # Convert Ltp to numeric by removing commas
            logging.info("Converting Ltp values to numeric format")
            df['Ltp_numeric'] = df['Ltp'].apply(convert_to_numeric)
            logging.debug(f"First few Ltp values converted: {list(zip(df['Ltp'].head(), df['Ltp_numeric'].head()))}")
            
            # Calculate MACD components
            logging.info("Calculating MACD components")
            df['MACD_Line'], df['Signal_Line'], df['Histogram'] = calculate_macd(df)
            
            # Check for MACD crossover
            logging.info("Checking for MACD crossover")
            macd_crosses_signal = check_macd_crossover(df, lookback_days=3)
            logging.info(f"MACD crossover result: {macd_crosses_signal}")
            
            # Create chart for this stock
            chart_path = create_macd_chart(df, sheet_name, output_dir)
            
            # Add results to the results worksheet
            logging.info("Adding results to worksheet")
            cell_stock = results_ws.cell(row=row, column=1, value=sheet_name)
            apply_cell_style(cell_stock)
            
            # Set Yes/No value with styles
            cell_macd_signal = results_ws.cell(row=row, column=2, value="Yes" if macd_crosses_signal else "No")
            if macd_crosses_signal:
                apply_yes_style(cell_macd_signal)
            else:
                apply_no_style(cell_macd_signal)
            
            # Add chart link placeholder
            cell_chart = results_ws.cell(row=row, column=3, value="View Chart")
            apply_cell_style(cell_chart)
            
            row += 1
            successful_sheets += 1
            logging.info(f"Successfully processed sheet: {sheet_name}")
            
        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {str(e)}")
            failed_sheets += 1
            continue
    
    # Add alternating row colors for readability
    for row_idx in range(5, row):
        if row_idx % 2 == 0:  # Even rows
            for col in range(1, 4):
                cell = results_ws.cell(row=row_idx, column=col)
                if not cell.fill or cell.fill.fill_type != "solid":
                    cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    
    # Add summary information
    summary_cell = results_ws.cell(row=3, column=1, value=f"Total Stocks Analyzed: {row-5}")
    summary_cell.font = Font(bold=True)
    results_ws.merge_cells('A3:C3')
    
    # Freeze panes for better navigation
    results_ws.freeze_panes = results_ws['A5']
    
    # Save the results workbook
    logging.info(f"Saving results workbook to {results_file}")
    results_wb.save(results_file)
    
    # Add hyperlinks to charts and navigation buttons
    logging.info("Adding hyperlinks to charts and navigation buttons")
    add_chart_hyperlinks_and_navigation(results_file, output_dir)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Analysis complete in {elapsed_time:.2f} seconds")
    logging.info(f"Successfully processed {successful_sheets} sheets, failed to process {failed_sheets} sheets")
    logging.info(f"Results saved to {results_file}")

def add_chart_hyperlinks_and_navigation(results_file, output_dir):
    """Add hyperlinks to chart images and navigation buttons in the Excel file"""
    logging.info("Adding chart hyperlinks and navigation buttons to the Excel file")
    wb = load_workbook(results_file)
    main_sheet = wb.active
    
    # Get the maximum row
    max_row = main_sheet.max_row
    logging.info(f"Processing {max_row-4} stock entries")  # Subtract header rows
    
    # Create a sheet for each chart
    for row in range(5, max_row + 1):  # Start from row 5 (data starts here)
        stock_name = main_sheet[f'A{row}'].value
        if not stock_name:
            continue
            
        chart_path = os.path.join(output_dir, "charts", f"{stock_name}_macd_chart.png")
        
        if os.path.exists(chart_path):
            logging.info(f"Creating chart sheet for {stock_name}")
            # Create a new sheet for this stock's chart
            chart_sheet_name = f"{stock_name}_Chart"
            # Use truncated name if too long (Excel has a 31 character limit for sheet names)
            if len(chart_sheet_name) > 31:
                chart_sheet_name = chart_sheet_name[:28] + "..."
                
            chart_sheet = wb.create_sheet(title=chart_sheet_name)
            
            # Add a title to the chart sheet
            chart_sheet['A1'] = f"{stock_name} - MACD Crossover Analysis"
            chart_sheet['A1'].font = Font(bold=True, size=14)
            chart_sheet['A1'].alignment = Alignment(horizontal="center")
            
            # Merge cells for the title
            chart_sheet.merge_cells('A1:F1')
            
            # Add the chart image to the sheet
            logging.info(f"Adding chart image to sheet: {chart_sheet_name}")
            img = Image(chart_path)
            
            # Set image dimensions to maintain correct aspect ratio (width:height = 10:7)
            img.width = 600  
            img.height = 420  # Maintain 10:7 ratio
            
            chart_sheet.add_image(img, 'A3')
            
            # Add "Go to Main Page" button at position H15 (adjusted for better placement)
            button_cell = chart_sheet['H15'] = "Go to Main Page"
            apply_button_style(chart_sheet['H15'])
            chart_sheet['H15'].hyperlink = f"#'Analysis Results'![A1"
            
            # Make sure column H is wide enough for button text
            chart_sheet.column_dimensions['H'].width = 20
            
            # Create a hyperlink from the results sheet to the chart sheet
            logging.info(f"Creating hyperlink for {stock_name}")
            link_cell = main_sheet[f'C{row}']
            link_cell.value = "View Chart"
            link_cell.hyperlink = f"#{chart_sheet_name}!A1"
            apply_hyperlink_style(link_cell)
        else:
            logging.warning(f"Chart not found for {stock_name}: {chart_path}")
    
    logging.info(f"Saving workbook with hyperlinks and navigation to {results_file}")
    wb.save(results_file)
    logging.info("Chart hyperlinks and navigation buttons added successfully")

if __name__ == "__main__":
    logging.info("===== MACD ANALYSIS SCRIPT STARTED =====")
    
    # Find the latest data file
    latest_file = find_latest_data_file()
    
    if latest_file:
        logging.info(f"Processing file: {latest_file}")
        generate_analysis_report(latest_file)
    else:
        logging.error("No data files found matching the pattern 'shares_data_till_*.xlsx'")
    
    logging.info("===== MACD ANALYSIS SCRIPT COMPLETED =====")