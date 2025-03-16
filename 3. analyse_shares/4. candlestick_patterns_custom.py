import numpy as np
import pandas as pd
import talib
import os
import glob
from datetime import datetime
import matplotlib.pyplot as plt
import mplfinance as mpf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import time

# Set matplotlib to use 'Agg' backend instead of tkinter to avoid threading issues
import matplotlib
matplotlib.use('Agg')  # This must be done before importing pyplot

def get_latest_data_file(directory, pattern="shares_data_till_*.xlsx"):
    """
    Find the most recent file matching the pattern in the specified directory
    """
    # Get all matching files
    matching_files = glob.glob(os.path.join(directory, pattern))
    
    if not matching_files:
        raise FileNotFoundError(f"No files matching '{pattern}' found in '{directory}'")
    
    # Find the most recent file by modified date
    latest_file = max(matching_files, key=os.path.getmtime)
    return latest_file

def read_stock_data_from_excel(file_path):
    """
    Read stock data from an Excel file with multiple sheets
    Returns a dictionary with ticker as key and dataframe as value
    """
    # Read all sheets from Excel file
    excel_data = pd.read_excel(file_path, sheet_name=None)
    
    stock_data = {}
    for sheet_name, df in excel_data.items():
        # Assuming sheet name is the ticker symbol
        ticker = sheet_name
        
        # Rename columns to match yfinance format
        df = df.rename(columns={
            'Date': 'Date',
            'Open': 'Open',
            'High': 'High',
            'Low': 'Low',
            'Ltp': 'Close',
            '% Change': 'Change',
            'Qty': 'Volume',
            'Turnover': 'Turnover'
        })
        
        # Convert comma-separated numbers to float for all numeric columns
        numeric_columns = ['Open', 'High', 'Low', 'Close', 'Volume', 'Turnover']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(',', '').astype(float)
        
        # Also convert Change column if it exists (might be a percentage)
        if 'Change' in df.columns:
            df['Change'] = df['Change'].astype(str).str.replace(',', '').astype(float)
        
        # Ensure Date is datetime
        df['Date'] = pd.to_datetime(df['Date'])
        df.set_index('Date', inplace=True)
        
        stock_data[ticker] = df
    
    return stock_data

def custom_bullish_engulfing(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average"):
    result = [None] * len(open_data)
    if len(open_data) < 2:
        return result
    
    for i in range(lookback, len(open_data)):
        # Check if first candle is bearish
        first_bearish = close_data[i-1] < open_data[i-1]
        # Check if second candle is bullish
        second_bullish = close_data[i] > open_data[i]
        # Check body engulfing
        body_engulfing = (open_data[i] <= close_data[i-1] and close_data[i] >= open_data[i-1])
        # Minimum body size ratio (second candle body > 1.5x first candle body)
        first_body = abs(open_data[i-1] - close_data[i-1])
        second_body = abs(open_data[i] - close_data[i])
        size_check = second_body > 1.5 * first_body
        
        # Volume condition
        volume_condition = False
        if volume_rule == "simple":
            # Simple rule: Volume of second candle > volume of first candle
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            # Average rule: Volume of second candle > average volume over lookback period
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume
        
        # Combine all conditions
        if first_bearish and second_bullish and body_engulfing and size_check and volume_condition:
            result[i] = "BULLISH ENGULFING"
    
    return result

def custom_bearish_engulfing(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average"):
    """
    Custom bearish engulfing with volume confirmation.
    Volume rule: 'simple' (second > first) or 'average' (second > lookback avg).
    """
    result = [None] * len(open_data)
    if len(open_data) < 2:
        return result
    
    for i in range(lookback, len(open_data)):
        # Check if first candle is bullish
        first_bullish = close_data[i-1] > open_data[i-1]
        # Check if second candle is bearish
        second_bearish = close_data[i] < open_data[i]
        # Check body engulfing
        body_engulfing = (open_data[i] >= close_data[i-1] and close_data[i] <= open_data[i-1])
        # Minimum body size ratio
        first_body = abs(close_data[i-1] - open_data[i-1])
        second_body = abs(open_data[i] - close_data[i])
        size_check = second_body > 1.5 * first_body
        
        # Volume condition (applied to second candle)
        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume
        
        if first_bullish and second_bearish and body_engulfing and size_check and volume_condition:
            result[i] = "BEARISH ENGULFING"
    
    return result

# Helper functions for trend conditions
def calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance=0.05, strict_trend=True):
    if i < lookback:
        return True  # Not enough data to determine trend, assume valid
    sma_slope = (sma[i] - sma[i - lookback]) / lookback
    if strict_trend:
        return sma_slope < 0 and abs(sma_slope) > trend_tolerance * abs(sma[i])  # Downtrend
    else:
        return sma_slope <= 0  # Downtrend or flat

def calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance=0.05, strict_trend=True):
    if i < lookback:
        return True  # Not enough data to determine trend, assume valid
    sma_slope = (sma[i] - sma[i - lookback]) / lookback
    if strict_trend:
        return sma_slope > 0 and abs(sma_slope) > trend_tolerance * abs(sma[i])  # Uptrend
    else:
        return sma_slope >= 0  # Uptrend or flat

# Candlestick pattern functions
def custom_morning_star(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(3, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data) - 2):
        # Day 1: Bearish candle
        day1_bearish = close_data[i-2] < open_data[i-2]
        day1_body = abs(open_data[i-2] - close_data[i-2])
        # Day 2: Small body, gapping down
        day2_body = abs(open_data[i-1] - close_data[i-1])
        small_body = day2_body < 0.3 * day1_body
        gap_down = open_data[i-1] < close_data[i-2]
        # Day 3: Bullish candle, closes above Day 1 midpoint
        day3_bullish = close_data[i] > open_data[i]
        day3_body = abs(close_data[i] - open_data[i])
        close_above_mid = close_data[i] > (open_data[i-2] + close_data[i-2]) / 2
        shadow_check = (high_data[i] - max(open_data[i], close_data[i])) < 0.2 * day3_body

        # Volume condition (applied to Day 3)
        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        # Prior trend: Downtrend
        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if (day1_bearish and small_body and gap_down and day3_bullish and close_above_mid and 
            shadow_check and volume_condition and trend_condition):
            result[i] = "MORNING STAR"

    return result

def custom_evening_star(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(3, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data) - 2):
        # Day 1: Bullish candle
        day1_bullish = close_data[i-2] > open_data[i-2]
        day1_body = abs(open_data[i-2] - close_data[i-2])
        # Day 2: Small body, gapping up
        day2_body = abs(open_data[i-1] - close_data[i-1])
        small_body = day2_body < 0.3 * day1_body
        gap_up = open_data[i-1] > close_data[i-2]
        # Day 3: Bearish candle, closes below Day 1 midpoint
        day3_bearish = close_data[i] < open_data[i]
        day3_body = abs(close_data[i] - open_data[i])
        close_below_mid = close_data[i] < (open_data[i-2] + close_data[i-2]) / 2
        shadow_check = (min(open_data[i], close_data[i]) - low_data[i]) < 0.2 * day3_body

        # Volume condition (applied to Day 3)
        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        # Prior trend: Uptrend
        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if (day1_bullish and small_body and gap_up and day3_bearish and close_below_mid and 
            shadow_check and volume_condition and trend_condition):
            result[i] = "EVENING STAR"

    return result

def custom_hammer(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        body = abs(close_data[i] - open_data[i])
        lower_shadow = min(open_data[i], close_data[i]) - low_data[i]
        upper_shadow = high_data[i] - max(open_data[i], close_data[i])
        
        small_body = body > 0
        long_lower_shadow = lower_shadow > 2.5 * body
        small_upper_shadow = upper_shadow < 0.3 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if small_body and long_lower_shadow and small_upper_shadow and volume_condition and trend_condition:
            result[i] = "HAMMER"

    return result

def custom_shooting_star(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        body = abs(close_data[i] - open_data[i])
        upper_shadow = high_data[i] - max(open_data[i], close_data[i])
        lower_shadow = min(open_data[i], close_data[i]) - low_data[i]
        
        small_body = body > 0
        long_upper_shadow = upper_shadow > 2.5 * body
        small_lower_shadow = lower_shadow < 0.3 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if small_body and long_upper_shadow and small_lower_shadow and volume_condition and trend_condition:
            result[i] = "SHOOTING STAR"

    return result

def custom_hanging_man(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        body = abs(close_data[i] - open_data[i])
        lower_shadow = min(open_data[i], close_data[i]) - low_data[i]
        upper_shadow = high_data[i] - max(open_data[i], close_data[i])
        
        small_body = body > 0
        long_lower_shadow = lower_shadow > 2.5 * body
        small_upper_shadow = upper_shadow < 0.3 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if small_body and long_lower_shadow and small_upper_shadow and volume_condition and trend_condition:
            result[i] = "HANGING MAN"

    return result

def custom_inverted_hammer(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        body = abs(close_data[i] - open_data[i])
        upper_shadow = high_data[i] - max(open_data[i], close_data[i])
        lower_shadow = min(open_data[i], close_data[i]) - low_data[i]
        
        small_body = body > 0
        long_upper_shadow = upper_shadow > 2.5 * body
        small_lower_shadow = lower_shadow < 0.3 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if small_body and long_upper_shadow and small_lower_shadow and volume_condition and trend_condition:
            result[i] = "INVERTED HAMMER"

    return result

def custom_piercing_pattern(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        if i < 1:
            continue
        
        bearish_day1 = close_data[i-1] < open_data[i-1]
        bullish_day2 = close_data[i] > open_data[i]
        opens_below = open_data[i] < low_data[i-1]
        midpoint_day1 = (open_data[i-1] + close_data[i-1]) / 2
        closes_above_mid = close_data[i] > midpoint_day1 and close_data[i] < open_data[i-1]

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if bearish_day1 and bullish_day2 and opens_below and closes_above_mid and volume_condition and trend_condition:
            result[i] = "PIERCING PATTERN"

    return result

def custom_dark_cloud_cover(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        if i < 1:
            continue
        
        bullish_day1 = close_data[i-1] > open_data[i-1]
        bearish_day2 = close_data[i] < open_data[i]
        opens_above = open_data[i] > high_data[i-1]
        midpoint_day1 = (open_data[i-1] + close_data[i-1]) / 2
        closes_below_mid = close_data[i] < midpoint_day1 and close_data[i] > open_data[i-1]

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if bullish_day1 and bearish_day2 and opens_above and closes_below_mid and volume_condition and trend_condition:
            result[i] = "DARK CLOUD COVER"

    return result

def custom_bullish_marubozu(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        bullish = close_data[i] > open_data[i]
        body = close_data[i] - open_data[i]
        upper_shadow = high_data[i] - close_data[i]
        lower_shadow = open_data[i] - low_data[i]
        small_shadows = upper_shadow < 0.1 * body and lower_shadow < 0.1 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if bullish and small_shadows and volume_condition and trend_condition:
            result[i] = "BULLISH MARUBOZU"

    return result

def custom_bearish_marubozu(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(2, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        bearish = close_data[i] < open_data[i]
        body = open_data[i] - close_data[i]
        upper_shadow = high_data[i] - open_data[i]
        lower_shadow = close_data[i] - low_data[i]
        small_shadows = upper_shadow < 0.1 * body and lower_shadow < 0.1 * body

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if bearish and small_shadows and volume_condition and trend_condition:
            result[i] = "BEARISH MARUBOZU"

    return result

def custom_rising_three_methods(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(5, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        if i < 4:
            continue
        
        day1_bullish = close_data[i-4] > open_data[i-4]
        day2_small = close_data[i-3] < open_data[i-3] and close_data[i-3] > open_data[i-4] and open_data[i-3] < close_data[i-4]
        day3_small = close_data[i-2] < open_data[i-2] and close_data[i-2] > open_data[i-4] and open_data[i-2] < close_data[i-4]
        day4_small = close_data[i-1] < open_data[i-1] and close_data[i-1] > open_data[i-4] and open_data[i-1] < close_data[i-4]
        day5_bullish = close_data[i] > open_data[i] and close_data[i] > close_data[i-4]

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_uptrend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if day1_bullish and day2_small and day3_small and day4_small and day5_bullish and volume_condition and trend_condition:
            result[i] = "RISING THREE METHODS"

    return result

def custom_falling_three_methods(open_data, close_data, high_data, low_data, volume_data, lookback=5, volume_rule="average", trend_tolerance=0.05, strict_trend=False):
    result = [None] * len(open_data)
    if len(open_data) < max(5, lookback):
        return result

    sma = talib.SMA(close_data, timeperiod=lookback)

    for i in range(lookback, len(open_data)):
        if i < 4:
            continue
        
        day1_bearish = close_data[i-4] < open_data[i-4]
        day2_small = close_data[i-3] > open_data[i-3] and close_data[i-3] < open_data[i-4] and open_data[i-3] > close_data[i-4]
        day3_small = close_data[i-2] > open_data[i-2] and close_data[i-2] < open_data[i-4] and open_data[i-2] > close_data[i-4]
        day4_small = close_data[i-1] > open_data[i-1] and close_data[i-1] < open_data[i-4] and open_data[i-1] > close_data[i-4]
        day5_bearish = close_data[i] < open_data[i] and close_data[i] < close_data[i-4]

        volume_condition = False
        if volume_rule == "simple":
            volume_condition = volume_data[i] > volume_data[i-1]
        elif volume_rule == "average":
            avg_volume = sum(volume_data[i-lookback:i]) / lookback
            volume_condition = volume_data[i] > avg_volume

        trend_condition = calculate_trend_condition(close_data, sma, i, lookback, trend_tolerance, strict_trend)

        if day1_bearish and day2_small and day3_small and day4_small and day5_bearish and volume_condition and trend_condition:
            result[i] = "FALLING THREE METHODS"

    return result


def detect_candlestick_patterns(data):
    """
    Use custom functions and TA-Lib to detect candlestick patterns with stricter rules
    """
    patterns = {}
    
    # Custom implementations
    patterns['BULLISH ENGULFING'] = custom_bullish_engulfing(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['BEARISH ENGULFING'] = custom_bearish_engulfing(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['MORNING STAR'] = custom_morning_star(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values,data['Volume'].values
    )
    patterns['EVENING STAR'] = custom_evening_star(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['HAMMER'] = custom_hammer(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['SHOOTING STAR'] = custom_shooting_star(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['HANGING MAN'] = custom_hanging_man(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    patterns['INVERTED HAMMER'] = custom_inverted_hammer(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )
    
    patterns['PIERCING PATTERN'] = custom_piercing_pattern(
    data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
)

    patterns['DARK CLOUD COVER'] = custom_dark_cloud_cover(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )

    patterns['BULLISH MARUBOZU'] = custom_bullish_marubozu(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )

    patterns['BEARISH MARUBOZU'] = custom_bearish_marubozu(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )

    patterns['RISING THREE METHODS'] = custom_rising_three_methods(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )

    patterns['FALLING THREE METHODS'] = custom_falling_three_methods(
        data['Open'].values, data['Close'].values, data['High'].values, data['Low'].values, data['Volume'].values
    )

    # TA-Lib for remaining patterns
    pattern_functions = {
    }
    
    for pattern_name, pattern_func in pattern_functions.items():
        if pattern_name in ['BULLISH MARUBOZU', 'BEARISH MARUBOZU']:
            result = pattern_func(data['Open'].values, data['High'].values, 
                                 data['Low'].values, data['Close'].values)
            if pattern_name == 'BULLISH MARUBOZU':
                patterns[pattern_name] = [pattern_name if x > 0 else None for x in result]
            else:
                patterns[pattern_name] = [pattern_name if x < 0 else None for x in result]
        elif pattern_name in ['RISING THREE METHODS', 'FALLING THREE METHODS']:
            result = pattern_func(data['Open'].values, data['High'].values, 
                                 data['Low'].values, data['Close'].values)
            if pattern_name == 'RISING THREE METHODS':
                patterns[pattern_name] = [pattern_name if x > 0 else None for x in result]
            else:
                patterns[pattern_name] = [pattern_name if x < 0 else None for x in result]
        else:
            result = pattern_func(data['Open'].values, data['High'].values, 
                                 data['Low'].values, data['Close'].values)
            patterns[pattern_name] = [pattern_name if x != 0 else None for x in result]
    
    return pd.DataFrame(patterns, index=data.index)

def analyze_stock_data(stock_data):
    """
    Analyze all stocks in the stock_data dictionary
    and return a dictionary of results with patterns and summary
    """
    results = {}
    total_stocks = len(stock_data)
    progress_step = max(1, total_stocks // 10)  # Show progress every 10%
    
    print(f"Analyzing {total_stocks} stocks...")
    
    for i, (ticker, data) in enumerate(stock_data.items(), 1):
        try:
            # Show progress
            if i % progress_step == 0 or i == total_stocks:
                print(f"Progress: {i}/{total_stocks} stocks analyzed ({i/total_stocks*100:.1f}%)")
            
            # Detect patterns
            patterns = detect_candlestick_patterns(data)
            
            # Get only the last 3trading days (or all days if less than 3)
            last_days = min(3,len(patterns))
            recent_patterns = patterns.tail(last_days)
            
            # Create a summary of detected patterns for the last days
            daily_summary = {}
            for day in recent_patterns.index:
                day_patterns = [p for p in recent_patterns.loc[day].values if p is not None]
                daily_summary[day.strftime('%Y-%m-%d')] = day_patterns
            
            # Categorize patterns into the four categories for summary page
            pattern_categories = {
                'bullish_reversal': ['BULLISH ENGULFING', 'MORNING STAR', 'HAMMER', 'PIERCING PATTERN','INVERTED HAMMER'],
                'bearish_reversal': ['BEARISH ENGULFING', 'EVENING STAR', 'SHOOTING STAR', 'DARK CLOUD COVER', 'HANGING MAN'],
                'bullish_continuation': ['BULLISH MARUBOZU', 'RISING THREE METHODS'],
                'bearish_continuation': ['BEARISH MARUBOZU', 'FALLING THREE METHODS']
            }
            
            # Find all unique patterns in recent days
            all_recent_patterns = set()
            for day_patterns in daily_summary.values():
                all_recent_patterns.update(day_patterns)
            
            # Categorize them
            categorized_patterns = {
                'bullish_reversal': [],
                'bearish_reversal': [],
                'bullish_continuation': [],
                'bearish_continuation': []
            }
            
            for pattern in all_recent_patterns:
                for category, category_patterns in pattern_categories.items():
                    if pattern in category_patterns:
                        categorized_patterns[category].append(pattern)
            
            # Store results
            results[ticker] = {
                'data': data,
                'patterns': patterns,
                'daily_summary': daily_summary,
                'categorized_patterns': categorized_patterns
            }
        except Exception as e:
            print(f"Error analyzing {ticker}: {str(e)}")
            # Skip this stock and continue with others
            continue
    
    return results

def create_candlestick_chart(stock_data, ticker, output_path):
    """
    Create and save a candlestick chart for a stock using mplfinance
    Returns the path to the saved chart
    """
    try:
        # Ensure directory exists and is writable
        chart_dir = os.path.dirname(output_path)
        os.makedirs(chart_dir, exist_ok=True)
        if not os.access(chart_dir, os.W_OK):
            raise PermissionError(f"No write permission for directory: {chart_dir}")
        
        # Create a copy for plotting
        plot_data = stock_data.copy()
        
        # Get the last 30 days of data or all if less than 30
        last_days = min(30, len(plot_data))
        plot_data = plot_data.tail(last_days)
        
        # Create the candlestick chart using mplfinance with improved styling
        plt.close('all')  # Clear any existing figures to prevent memory leaks
        mpf.plot(plot_data, type='candle', style='yahoo', 
                 title=f'{ticker} Candlestick Chart (Last {last_days} Trading Days)',
                 ylabel='Price', volume=True, figsize=(12, 8),
                 savefig=dict(fname=output_path, dpi=150),
                 show_nontrading=False)
        
        # Explicitly close the figure to free up memory
        plt.close('all')
        
        # Verify file was created
        if not os.path.exists(output_path):
            print(f"Warning: Chart file was not created for {ticker} at {output_path}")
            return None
        
        return output_path
    except Exception as e:
        print(f"Error creating chart for {ticker}: {str(e)}")
        return None

def create_excel_report(results, output_dir):
    """
    Create an enhanced Excel report with a summary page and individual stock pages,
    ensuring scrollability on chart pages.
    """
    # Create output directories if they don't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create a timestamp for the filename
    timestamp = datetime.now().strftime('%Y_%m_%d')
    output_file = os.path.join(output_dir, f'candlestick_patterns_{timestamp}.xlsx')
    
    # Create a new workbook
    wb = Workbook()
    
    # Create the main summary sheet
    summary_sheet = wb.active
    summary_sheet.title = 'Summary'
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True, color='0000FF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    no_pattern_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Add title
    title_cell = summary_sheet.cell(row=1, column=1, value="Candlestick Pattern Analysis Report")
    title_cell.font = title_font
    summary_sheet.merge_cells('A1:F1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add generated date
    date_cell = summary_sheet.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.font = Font(italic=True)
    
    # Add headers
    headers = ['Stock', 'Bullish Reversal', 'Bearish Reversal', 'Bullish Continuation', 
               'Bearish Continuation', 'Go to Chart']
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Fill in data for each stock
    for row, (ticker, result) in enumerate(results.items(), header_row + 1):
        stock_cell = summary_sheet.cell(row=row, column=1, value=ticker)
        stock_cell.border = border
        stock_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        bullish_reversal = ', '.join(result['categorized_patterns']['bullish_reversal'])
        cell = summary_sheet.cell(row=row, column=2, value=bullish_reversal if bullish_reversal else "")
        cell.fill = green_fill if bullish_reversal else no_pattern_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        bearish_reversal = ', '.join(result['categorized_patterns']['bearish_reversal'])
        cell = summary_sheet.cell(row=row, column=3, value=bearish_reversal if bearish_reversal else "")
        cell.fill = red_fill if bearish_reversal else no_pattern_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        bullish_continuation = ', '.join(result['categorized_patterns']['bullish_continuation'])
        cell = summary_sheet.cell(row=row, column=4, value=bullish_continuation if bullish_continuation else "")
        cell.fill = green_fill if bullish_continuation else no_pattern_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        bearish_continuation = ', '.join(result['categorized_patterns']['bearish_continuation'])
        cell = summary_sheet.cell(row=row, column=5, value=bearish_continuation if bearish_continuation else "")
        cell.fill = red_fill if bearish_continuation else no_pattern_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        link_cell = summary_sheet.cell(row=row, column=6, value='View Chart')
        link_cell.hyperlink = f"#{ticker}!A1"
        link_cell.style = "Hyperlink"
        link_cell.border = border
        link_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add a legend
    legend_row = summary_sheet.max_row + 3
    summary_sheet.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    
    bullish_cell = summary_sheet.cell(row=legend_row + 1, column=1, value="Bullish Patterns")
    bullish_cell.fill = green_fill
    bullish_cell.border = border
    
    bearish_cell = summary_sheet.cell(row=legend_row + 2, column=1, value="Bearish Patterns")
    bearish_cell.fill = red_fill
    bearish_cell.border = border
    
    none_cell = summary_sheet.cell(row=legend_row + 3, column=1, value="No Patterns")
    none_cell.fill = no_pattern_fill
    none_cell.border = border
    
    # Auto-adjust column widths on summary sheet
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for r in range(1, summary_sheet.max_row + 1):
            cell_value = summary_sheet.cell(row=r, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max(max_length, len(headers[col-1])) + 2
        summary_sheet.column_dimensions[column_letter].width = min(adjusted_width, 40)
    
    # Create a temp directory for charts
    temp_chart_dir = os.path.join(output_dir, 'temp_charts')
    os.makedirs(temp_chart_dir, exist_ok=True)
    chart_files = []
    
    # Process stocks in batches
    total_stocks = len(results)
    batch_size = 20
    print(f"\nCreating report sheets for {total_stocks} stocks in batches of {batch_size}...")
    
    ticker_list = list(results.keys())
    for batch_start in range(0, total_stocks, batch_size):
        batch_count = batch_start // batch_size + 1
        batch_end = min(batch_start + batch_size, total_stocks)
        batch_tickers = ticker_list[batch_start:batch_end]
        
        print(f"Processing batch {batch_count}: stocks {batch_start+1}-{batch_end} of {total_stocks}")
        
        for ticker in batch_tickers:
            result = results[ticker]
            stock_sheet = wb.create_sheet(title=ticker)
            
            # Ensure worksheet has enough rows for scrolling (at least 100 rows)
            stock_sheet.sheet_view.topLeftCell = 'A1'  # Reset view to top-left
            stock_sheet.sheet_view.showGridLines = True  # Ensure gridlines are visible
            
            # Add title
            title_cell = stock_sheet.cell(row=1, column=1, value=f"{ticker} Candlestick Analysis")
            title_cell.font = title_font
            stock_sheet.merge_cells('A1:E1')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add "Go to Main Page" button in cell H15
            main_page_cell = stock_sheet.cell(row=15, column=8, value="Go to Main Page")
            main_page_cell.hyperlink = "#Summary!A1"
            main_page_cell.style = "Hyperlink"
            main_page_cell.font = Font(bold=True, color="0000FF")
            main_page_cell.alignment = Alignment(horizontal='center', vertical='center')
            main_page_cell.border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            
            # Create and add candlestick chart
            chart_path = os.path.join(temp_chart_dir, f"{ticker}_chart.png")
            chart_result = create_candlestick_chart(result['data'], ticker, chart_path)
            
            if chart_result and os.path.exists(chart_path):
                img = Image(chart_path)
                img.width = 800  # Fixed width in pixels
                img.height = 500  # Fixed height in pixels
                stock_sheet.add_image(img, 'A3')  # Anchor at A3
                chart_files.append(chart_path)
            else:
                error_cell = stock_sheet.cell(row=3, column=1, value=f"Unable to create chart for {ticker}")
                error_cell.font = Font(color="FF0000")
            
            # Add pattern information below the chart
            pattern_row = 35  # Adjusted to ensure it’s below the chart
            pattern_title = stock_sheet.cell(row=pattern_row, column=1, value="Detected Patterns:")
            pattern_title.font = Font(bold=True, size=12)
            stock_sheet.merge_cells(f'A{pattern_row}:C{pattern_row}')
            
            header_row = pattern_row + 1
            headers = ["Date", "Detected Patterns", "Pattern Type"]
            for col, header in enumerate(headers, 1):
                cell = stock_sheet.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for i, (day, patterns) in enumerate(result['daily_summary'].items()):
                day_row = header_row + i + 1
                date_cell = stock_sheet.cell(row=day_row, column=1, value=day)
                date_cell.border = border
                date_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if patterns:
                    pattern_text = ', '.join(patterns)
                    pattern_cell = stock_sheet.cell(row=day_row, column=2, value=pattern_text)
                    pattern_types = []
                    for pattern in patterns:
                        if pattern in result['categorized_patterns']['bullish_reversal']:
                            pattern_types.append("Bullish Reversal")
                        elif pattern in result['categorized_patterns']['bearish_reversal']:
                            pattern_types.append("Bearish Reversal")
                        elif pattern in result['categorized_patterns']['bullish_continuation']:
                            pattern_types.append("Bullish Continuation")
                        elif pattern in result['categorized_patterns']['bearish_continuation']:
                            pattern_types.append("Bearish Continuation")
                    
                    if any(p_type.startswith("Bullish") for p_type in pattern_types):
                        pattern_cell.fill = green_fill
                    elif any(p_type.startswith("Bearish") for p_type in pattern_types):
                        pattern_cell.fill = red_fill
                    
                    type_cell = stock_sheet.cell(row=day_row, column=3, value=', '.join(pattern_types))
                    if any(p_type.startswith("Bullish") for p_type in pattern_types):
                        type_cell.fill = green_fill
                    elif any(p_type.startswith("Bearish") for p_type in pattern_types):
                        type_cell.fill = red_fill
                else:
                    pattern_cell = stock_sheet.cell(row=day_row, column=2, value="")
                    pattern_cell.fill = no_pattern_fill
                    type_cell = stock_sheet.cell(row=day_row, column=3, value="")
                    type_cell.fill = no_pattern_fill
                
                pattern_cell.border = border
                pattern_cell.alignment = Alignment(horizontal='left', vertical='center')
                type_cell.border = border
                type_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add explanation of patterns
            explanation_row = day_row + 3
            explanation_title = stock_sheet.cell(row=explanation_row, column=1, value="Pattern Explanations:")
            explanation_title.font = Font(bold=True, size=12)
            stock_sheet.merge_cells(f'A{explanation_row}:D{explanation_row}')
            
            explanations = {
                'BULLISH ENGULFING': "A bullish reversal pattern where a large bullish candle engulfs the body of the previous bearish candle, indicating a potential upward trend.",
                'BEARISH ENGULFING': "A bearish reversal pattern where a large bearish candle engulfs the body of the previous bullish candle, indicating a potential downward trend.",
                'MORNING STAR': "A bullish reversal pattern consisting of three candles: a bearish candle, a small-bodied candle, and a bullish candle, indicating a potential upward trend.",
                'EVENING STAR': "A bearish reversal pattern consisting of three candles: a bullish candle, a small-bodied candle, and a bearish candle, indicating a potential downward trend.",
                'HAMMER': "A bullish reversal pattern with a small body and a long lower shadow, indicating potential support and upward momentum.",
                'SHOOTING STAR': "A bearish reversal pattern with a small body and a long upper shadow, indicating potential resistance and downward momentum.",
                'PIERCING PATTERN': "A bullish reversal pattern where a bullish candle opens below the previous bearish candle's low and closes above its midpoint, indicating a potential upward trend.",
                'DARK CLOUD COVER': "A bearish reversal pattern where a bearish candle opens above the previous bullish candle's high and closes below its midpoint, indicating a potential downward trend.",
                'BULLISH MARUBOZU': "A bullish continuation pattern with no or very small shadows, indicating strong buying pressure.",
                'BEARISH MARUBOZU': "A bearish continuation pattern with no or very small shadows, indicating strong selling pressure.",
                'RISING THREE METHODS': "A bullish continuation pattern with a large bullish candle followed by three small bearish candles and another large bullish candle, indicating a continued upward trend.",
                'FALLING THREE METHODS': "A bearish continuation pattern with a large bearish candle followed by three small bullish candles and another large bearish candle, indicating a continued downward trend.",
                'INVERTED HAMMER': "A bullish reversal pattern with a small body at the bottom and a long upper shadow, appearing in downtrends and signaling potential reversal.",
                'HANGING MAN': "A bearish reversal pattern with a small body at the top and a long lower shadow, appearing in uptrends and signaling potential reversal."
            }
            
            all_patterns = set()
            for day_patterns in result['daily_summary'].values():
                all_patterns.update(day_patterns)
            
            for i, pattern in enumerate(sorted(all_patterns)):
                if pattern in explanations:
                    exp_row = explanation_row + i + 1
                    pattern_cell = stock_sheet.cell(row=exp_row, column=1, value=pattern)
                    pattern_cell.font = Font(bold=True)
                    
                    if pattern in result['categorized_patterns']['bullish_reversal'] or pattern in result['categorized_patterns']['bullish_continuation']:
                        pattern_cell.fill = green_fill
                    elif pattern in result['categorized_patterns']['bearish_reversal'] or pattern in result['categorized_patterns']['bearish_continuation']:
                        pattern_cell.fill = red_fill
                    
                    explanation_cell = stock_sheet.cell(row=exp_row, column=2, value=explanations[pattern])
                    explanation_cell.alignment = Alignment(wrap_text=True)
                    stock_sheet.merge_cells(f'B{exp_row}:D{exp_row}')
                    pattern_cell.border = border
                    explanation_cell.border = border
            
            # Set column widths and ensure enough rows for scrolling
            for col in range(1, 10):
                column_letter = get_column_letter(col)
                max_length = 0
                for r in range(1, stock_sheet.max_row + 1):
                    cell_value = stock_sheet.cell(row=r, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = max(max_length + 2, 15)
                stock_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            # Ensure worksheet has enough rows (at least 100) to enable scrolling
            if stock_sheet.max_row < 100:
                stock_sheet.cell(row=100, column=1, value="")  # Dummy cell to extend sheet
            
        # Save after each batch
        try:
            wb.save(output_file)
            time.sleep(1)
            print(f"Progress saved after batch {batch_count}")
        except Exception as e:
            print(f"Warning: Could not save intermediate progress for batch {batch_count}: {str(e)}")
    
    # Save the final workbook
    print(f"Saving Excel report to: {output_file}")
    try:
        wb.save(output_file)
    except Exception as e:
        print(f"Error saving final report: {str(e)}")
        alt_output = os.path.join(os.path.dirname(output_dir), f'emergency_candlestick_patterns_{timestamp}.xlsx')
        print(f"Attempting to save to alternative location: {alt_output}")
        wb.save(alt_output)
        output_file = alt_output
    
    # Clean up temporary chart files
    print("Cleaning up temporary files...")
    for chart_file in chart_files:
        try:
            os.remove(chart_file)
        except Exception as e:
            print(f"Warning: Could not remove temp file {chart_file}: {str(e)}")
    try:
        os.rmdir(temp_chart_dir)
    except Exception as e:
        print(f"Warning: Could not remove temp directory: {str(e)}")
    
    print(f"Report successfully saved to: {output_file}")
    return output_file

def main():
    # Get current script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    try:
        # Find the latest data file
        print("Looking for the latest shares data file...")
        input_file = get_latest_data_file(script_dir, "shares_data_till_*.xlsx")
        print(f"Found latest data file: {os.path.basename(input_file)}")
        
        # Define base output directory (one level up in the specified folder)
        base_output_dir = os.path.join(os.path.dirname(script_dir), '4. analysis_result', 'candlestick pattern')
        
        # Create a subfolder with today's date in yyyy-mm-dd format
        today_date = datetime.now().strftime('%Y-%m-%d')  # e.g., 2025-03-13
        output_dir = os.path.join(base_output_dir, today_date)
        
        # Create the output directory structure if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        print(f"Reading stock data from: {input_file}")
        start_time = time.time()
        
        # Read stock data from Excel
        stock_data = read_stock_data_from_excel(input_file)
        print(f"Found {len(stock_data)} stocks in the Excel file.")
        
        # Analyze the stock data
        print("Starting analysis...")
        results = analyze_stock_data(stock_data)
        
        # More accurate progress reporting
        print("Analysis in progress...")
        time.sleep(1)  # Small delay to avoid reporting completion before it's done
        
        # Create the Excel report
        print("Creating Excel report...")
        output_file = create_excel_report(results, output_dir)
        
        # Report completion time
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"Analysis and report creation completed in {elapsed_time:.2f} seconds.")
        print(f"Final report saved to: {output_file}")
        
    except FileNotFoundError as e:
        print(f"Error: {str(e)}")
        print("Please make sure there are files matching 'shares_data_till_*.xlsx' in the script directory.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
