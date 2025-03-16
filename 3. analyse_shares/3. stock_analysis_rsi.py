import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import glob
import time
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Find the most recent data file
def find_latest_data_file():
    logging.info("Searching for the most recent data file...")
    files = glob.glob("shares_data_till_*.xlsx")
    
    if not files:
        logging.warning("No files found matching the pattern 'shares_data_till_*.xlsx'")
        return None
        
    logging.info(f"Found {len(files)} files matching the pattern")
    
    latest_file = None
    latest_date = None
    
    for file in files:
        try:
            # Extract date part from filename
            date_str = file.replace("shares_data_till_", "").replace(".xlsx", "")
            # Updated to handle yyyy_mm_dd format
            date_parts = date_str.split("_")
            logging.debug(f"Processing file: {file}, date parts: {date_parts}")
            
            if len(date_parts) == 3:  # Format is yyyy_mm_dd
                file_date = f"{date_parts[0]}-{date_parts[1]}-{date_parts[2]}"
                logging.debug(f"Extracted date: {file_date}")
                
                if latest_date is None or file_date > latest_date:
                    latest_date = file_date
                    latest_file = file
                    logging.debug(f"New latest file: {latest_file} with date {latest_date}")
        except Exception as e:
            logging.warning(f"Error processing filename {file}: {str(e)}")
            continue
    
    if latest_file:
        logging.info(f"Selected latest file: {latest_file} (date: {latest_date})")
    else:
        logging.warning("Could not determine the latest file")
        
    return latest_file


def convert_to_numeric(x):
    """Convert string values with commas to float"""
    if isinstance(x, str):
        return float(x.replace(',', ''))
    return x
def calculate_rsi(df, period=14):
    """Calculate Relative Strength Index (RSI) using Smoothed Moving Average"""
    logging.debug(f"Calculating {period}-period RSI using Smoothed Moving Average")
    
    # Ensure Ltp is numeric
    ltp = df['Ltp_numeric']
    
    # Calculate price changes
    delta = ltp.diff()
    
    # Separate gains and losses
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)
    
    # Calculate first average gain and loss (simple average for first period)
    first_avg_gain = gain.iloc[:period].mean()
    first_avg_loss = loss.iloc[:period].mean()
    
    # Initialize smoothed averages with first averages
    avg_gains = np.zeros(len(df))
    avg_losses = np.zeros(len(df))
    
    # Set the first period average
    avg_gains[period-1] = first_avg_gain
    avg_losses[period-1] = first_avg_loss
    
    # Calculate smoothed averages for the rest of the periods
    # Formula: SMA = (Previous SMA * (period-1) + Current Value) / period
    for i in range(period, len(df)):
        avg_gains[i] = (avg_gains[i-1] * (period-1) + gain.iloc[i]) / period
        avg_losses[i] = (avg_losses[i-1] * (period-1) + loss.iloc[i]) / period
    
    # Convert to pandas Series
    avg_gain_series = pd.Series(avg_gains, index=df.index)
    avg_loss_series = pd.Series(avg_losses, index=df.index)
    
    # Calculate RS (Relative Strength)
    rs = avg_gain_series / avg_loss_series
    
    # Calculate RSI
    rsi = 100 - (100 / (1 + rs))
    
    # First period values will be NaN
    rsi[:period] = np.nan
    
    return rsi

def check_rsi_conditions(df, lookback_days=3):
    """Check RSI conditions for recent days"""
    logging.debug("Checking RSI conditions")
    
    # Get recent data where RSI has values
    recent_data = df.dropna(subset=['RSI_14']).tail(lookback_days)
    
    if len(recent_data) < lookback_days:
        logging.warning(f"Not enough data points with valid RSI values. Got {len(recent_data)}, need {lookback_days}")
        return [], False, False, False
    
    # Get the latest RSI values
    latest_rsi_values = recent_data['RSI_14'].tolist()
    
    # Check if RSI is in uptrend territory (>60)
    is_uptrend = recent_data['RSI_14'].iloc[-1] > 60
    
    # Check if RSI is in sideways territory (between 40 and 60)
    is_sideways = 40 <= recent_data['RSI_14'].iloc[-1] <= 60
    
    # Check if RSI is in downtrend territory (<40)
    is_downtrend = recent_data['RSI_14'].iloc[-1] < 40
    
    logging.debug(f"Latest RSI values: {latest_rsi_values}")
    logging.debug(f"Uptrend: {is_uptrend}, Sideways: {is_sideways}, Downtrend: {is_downtrend}")
    
    return latest_rsi_values, is_uptrend, is_sideways, is_downtrend

def check_rsi_divergence(df, lookback_window=20):
    """Check for bullish and bearish RSI divergences"""
    logging.debug("Checking for RSI divergences")
    
    # Get recent data where RSI has values
    recent_data = df.dropna(subset=['RSI_14']).tail(lookback_window)
    
    if len(recent_data) < 10:  # Need a reasonable amount of data to check for divergences
        logging.warning("Not enough data points to check for divergences")
        return False, False
    
    # Simple implementation - this would be more sophisticated in a real trading system
    # Looking for local minima and maxima in price and RSI
    
    # Check for bullish divergence (price makes lower low but RSI makes higher low)
    # This is a simplified check - would be more complex in production
    price_min_idx = recent_data['Ltp_numeric'].idxmin()
    price_min = recent_data.loc[price_min_idx, 'Ltp_numeric']
    
    # Find a previous price minimum
    prev_data = df.loc[:price_min_idx].tail(lookback_window)
    if len(prev_data) > 5:
        prev_price_min_idx = prev_data['Ltp_numeric'].idxmin()
        prev_price_min = prev_data.loc[prev_price_min_idx, 'Ltp_numeric']
        
        # Check if price made lower low
        price_lower_low = price_min < prev_price_min
        
        # Check if RSI made higher low
        rsi_at_price_min = recent_data.loc[price_min_idx, 'RSI_14']
        rsi_at_prev_price_min = prev_data.loc[prev_price_min_idx, 'RSI_14']
        rsi_higher_low = rsi_at_price_min > rsi_at_prev_price_min
        
        bullish_divergence = price_lower_low and rsi_higher_low
    else:
        bullish_divergence = False
    
    # Check for bearish divergence (price makes higher high but RSI makes lower high)
    price_max_idx = recent_data['Ltp_numeric'].idxmax()
    price_max = recent_data.loc[price_max_idx, 'Ltp_numeric']
    
    # Find a previous price maximum
    prev_data = df.loc[:price_max_idx].tail(lookback_window)
    if len(prev_data) > 5:
        prev_price_max_idx = prev_data['Ltp_numeric'].idxmax()
        prev_price_max = prev_data.loc[prev_price_max_idx, 'Ltp_numeric']
        
        # Check if price made higher high
        price_higher_high = price_max > prev_price_max
        
        # Check if RSI made lower high
        rsi_at_price_max = recent_data.loc[price_max_idx, 'RSI_14']
        rsi_at_prev_price_max = prev_data.loc[prev_price_max_idx, 'RSI_14']
        rsi_lower_high = rsi_at_price_max < rsi_at_prev_price_max
        
        bearish_divergence = price_higher_high and rsi_lower_high
    else:
        bearish_divergence = False
    
    logging.debug(f"Bullish divergence: {bullish_divergence}, Bearish divergence: {bearish_divergence}")
    return bullish_divergence, bearish_divergence

def create_rsi_chart(df, sheet_name, output_dir):
    """Create a chart showing RSI and save it as an image"""
    logging.info(f"Creating RSI chart for {sheet_name}")
    
    # Create a figure with two subplots (price above, RSI below)
    # Change the height_ratios from [3, 1] to [2, 1.5] to give RSI more space
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10), gridspec_kw={'height_ratios': [2, 1.5]}, sharex=True)
    
    # Use a more attractive color scheme and styling
    plt.style.use('ggplot')
    
    # Plot price on the top subplot
    ax1.plot(df['Date'], df['Ltp_numeric'], label='Price', color='#404040', linewidth=1.5)
    ax1.set_title(f'{sheet_name} - Price and RSI Analysis', fontsize=16, fontweight='bold')
    ax1.set_ylabel('Price', fontsize=14)
    ax1.grid(True, alpha=0.3)
    ax1.legend(loc='best', fontsize=12)
    
    # Plot RSI on the bottom subplot
    ax2.plot(df['Date'], df['RSI_14'], label='RSI(14)', color='#1f77b4', linewidth=2.5)
    ax2.set_ylim(0, 100)
    ax2.set_ylabel('RSI', fontsize=14)
    ax2.set_xlabel('Date', fontsize=14)
    ax2.grid(True, alpha=0.3)
    ax2.tick_params(axis='both', which='major', labelsize=12)  # Increase tick label size
    
    # Add uptrend/downtrend levels with increased linewidth
    ax2.axhline(y=60, color='#d62728', linestyle='-', alpha=0.8, linewidth=2)
    ax2.axhline(y=40, color='#2ca02c', linestyle='-', alpha=0.8, linewidth=2)
    ax2.axhline(y=50, color='#7f7f7f', linestyle='--', alpha=0.6, linewidth=2)

    # Add text labels for the levels with increased font size
    ax2.text(df['Date'].iloc[0], 63, 'Uptrend (60)', fontsize=12, color='#d62728', fontweight='bold')
    ax2.text(df['Date'].iloc[0], 43, 'Downtrend (40)', fontsize=12, color='#2ca02c', fontweight='bold')
    ax2.text(df['Date'].iloc[0], 53, 'Midline (50)', fontsize=12, color='#7f7f7f', fontweight='bold')
    
    # Add some padding around the plot
    plt.tight_layout()
    
    # Create directory if it doesn't exist
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
    
    # Save chart as image with higher DPI for better quality
    chart_path = os.path.join(charts_dir, f"{sheet_name}_rsi_chart.png")
    plt.savefig(chart_path, dpi=150)  # Increased DPI for sharper image
    plt.close()
    logging.info(f"Chart saved to {chart_path}")
    
    return chart_path

def apply_header_style(cell):
    """Apply header style to a cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_button_style(cell):
    """Apply button style to a cell"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_uptrend_style(cell):
    """Apply uptrend (red) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_downtrend_style(cell):
    """Apply downtrend (green) style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_neutral_style(cell):
    """Apply neutral style to a cell"""
    cell.font = Font(bold=True, color="000000")
    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_cell_style(cell):
    """Apply basic cell style"""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_hyperlink_style(cell):
    """Apply hyperlink style to a cell"""
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def apply_note_style(cell):
    """Apply note style to a cell"""
    cell.font = Font(size=10, italic=True)
    cell.alignment = Alignment(vertical="center", wrap_text=True)

def generate_analysis_report(input_file):
    """Generate RSI analysis report for all sheets in the input file"""
    start_time = time.time()
    logging.info(f"Starting RSI analysis of file: {input_file}")
    
    # Create output directory structure with today's date
    base_output_dir = os.path.join(os.path.dirname(os.getcwd()), "4. analysis_result")
    rsi_dir = os.path.join(base_output_dir, "rsi")
    
    # Create date-based subfolder
    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.join(rsi_dir, today)
    
    # Create necessary directories
    if not os.path.exists(output_dir):
        logging.info(f"Creating output directory: {output_dir}")
        os.makedirs(output_dir)
    else:
        logging.info(f"Using existing output directory: {output_dir}")
    
    # Create charts directory
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        logging.info(f"Creating charts directory: {charts_dir}")
        os.makedirs(charts_dir)
    
    # Read the Excel file
    logging.info("Reading Excel file and getting sheet names")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    logging.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
    
    # Create a workbook for the analysis results
    results_file = os.path.join(output_dir, "rsi_analysis_results.xlsx")
    logging.info(f"Creating results workbook: {results_file}")
    results_wb = Workbook()
    results_ws = results_wb.active
    results_ws.title = "Analysis Results"
    
    # Add title and date to the main sheet
    title_cell = results_ws.cell(row=1, column=1, value="RSI Analysis")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    results_ws.merge_cells('A1:F1')
    
    # Add date information
    date_cell = results_ws.cell(row=2, column=1, value=f"Generated on: {today}")
    date_cell.font = Font(italic=True)
    
    # Add RSI usage notes
    notes = [
    "RSI Usage Guidelines:",
    "1. RSI above 60 indicates an uptrend. The market is showing strength and momentum.",
    "2. RSI between 40-60 indicates a sideways trend. The 40 and 60 levels often act as support or resistance.",
    "3. RSI below 40 indicates a downtrend. The market is showing weakness.",
    "4. Watch for RSI breaking through the 40 or 60 levels, as this can signal potential trend reversals.",
    "5. For long-term investment (monthly charts), RSI below 40 represents an undervalued region, providing potential buying opportunities.",
    "6. Bullish divergence: Price makes a lower low, but RSI makes a higher low (possible upward reversal).",
    "7. Bearish divergence: Price makes a higher high, but RSI makes a lower high (possible downward reversal).",
    "8. When RSI oscillates between 40-60 repeatedly, keep the stock on watchlist as it may be consolidating before a breakout.",
    "9. RSI can remain in uptrend/downtrend zones during strong market trends."
    ]
    
    # Add notes with proper formatting
    current_row = 3
    for note in notes:
        note_cell = results_ws.cell(row=current_row, column=1, value=note)
        apply_note_style(note_cell)
        results_ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1
    
    # Leave a blank row
    current_row += 1
    
    # Add headers to the results worksheet
    logging.info("Adding headers to results worksheet")
    headers = ["Stock", "Latest RSI", "Day-1 RSI", "Day-2 RSI", "RSI Status", "Chart Link"]
    for col, header in enumerate(headers, 1):
        cell = results_ws.cell(row=current_row, column=col, value=header)
        apply_header_style(cell)
    
    # Set column widths
    results_ws.column_dimensions['A'].width = 25
    results_ws.column_dimensions['B'].width = 15
    results_ws.column_dimensions['C'].width = 15
    results_ws.column_dimensions['D'].width = 15
    results_ws.column_dimensions['E'].width = 15
    results_ws.column_dimensions['F'].width = 15
    
    # Process each sheet
    data_row = current_row + 1  # Start data from next row
    successful_sheets = 0
    failed_sheets = 0
    created_charts = {}  # Dictionary to track successfully created charts
    
    for sheet_idx, sheet_name in enumerate(sheet_names):
        try:
            logging.info(f"Processing sheet {sheet_idx+1}/{len(sheet_names)}: {sheet_name}")
            
            # Read data from the sheet
            logging.info(f"Reading data from sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logging.info(f"Data shape: {df.shape} rows × {df.shape[1]} columns")
            
            # Convert Date column to datetime if needed
            logging.info("Converting Date column to datetime format")
            if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
                df['Date'] = pd.to_datetime(df['Date'])
            
            # Convert Ltp to numeric by removing commas
            logging.info("Converting Ltp values to numeric format")
            df['Ltp_numeric'] = df['Ltp'].apply(convert_to_numeric)
            logging.debug(f"First few Ltp values converted: {list(zip(df['Ltp'].head(), df['Ltp_numeric'].head()))}")
            
            # Calculate RSI
            logging.info("Calculating 14-period RSI")
            df['RSI_14'] = calculate_rsi(df, period=14)
            
            # Check RSI conditions
            logging.info("Checking RSI conditions")
            latest_rsi_values, is_uptrend, is_sideways, is_downtrend = check_rsi_conditions(df, lookback_days=3)

            # Check for RSI divergences
            logging.info("Checking RSI divergences")
            bullish_divergence, bearish_divergence = check_rsi_divergence(df)

            # Set RSI status
            if is_uptrend:
                rsi_status = "Uptrend"
            elif is_downtrend:
                rsi_status = "Downtrend"
            elif bullish_divergence:
                rsi_status = "Bullish Divergence"
            elif bearish_divergence:
                rsi_status = "Bearish Divergence"
            else:
                rsi_status = "Sideways"
                
            # Create chart for this stock
            try:
                # Check if we have enough valid data for plotting
                if len(df.dropna(subset=['RSI_14'])) >= 5:
                    logging.info(f"Creating chart for {sheet_name}")
                    chart_path = create_rsi_chart(df, sheet_name, output_dir)
                    if chart_path and os.path.exists(chart_path):
                        created_charts[sheet_name] = chart_path
                        logging.info(f"Chart created successfully for {sheet_name}")
                    else:
                        logging.warning(f"Failed to create chart for {sheet_name}")
                else:
                    logging.warning(f"Not enough valid RSI data for {sheet_name} to create chart")
            except Exception as e:
                logging.error(f"Error creating chart for {sheet_name}: {str(e)}")
                # Continue with the analysis even if chart creation fails
            
            # Add results to the results worksheet
            logging.info("Adding results to worksheet")
            
            # Stock name
            cell_stock = results_ws.cell(row=data_row, column=1, value=sheet_name)
            apply_cell_style(cell_stock)
            
            # Add latest RSI values for the last 3 days (if available)
            if len(latest_rsi_values) >= 3:
                # Latest RSI
                latest_rsi = latest_rsi_values[-1]
                cell_latest_rsi = results_ws.cell(row=data_row, column=2, value=round(latest_rsi, 2))
                apply_cell_style(cell_latest_rsi)
                
                # Day-1 RSI
                day1_rsi = latest_rsi_values[-2]
                cell_day1_rsi = results_ws.cell(row=data_row, column=3, value=round(day1_rsi, 2))
                apply_cell_style(cell_day1_rsi)
                
                # Day-2 RSI
                day2_rsi = latest_rsi_values[-3]
                cell_day2_rsi = results_ws.cell(row=data_row, column=4, value=round(day2_rsi, 2))
                apply_cell_style(cell_day2_rsi)
            else:
                # Handle cases where we don't have 3 days of data
                for i, col in enumerate(range(2, 5)):
                    if i < len(latest_rsi_values):
                        cell_rsi = results_ws.cell(row=data_row, column=col, value=round(latest_rsi_values[-(i+1)], 2))
                    else:
                        cell_rsi = results_ws.cell(row=data_row, column=col, value="N/A")
                    apply_cell_style(cell_rsi)
            
            # RSI Status with appropriate style
            cell_status = results_ws.cell(row=data_row, column=5, value=rsi_status)
            if rsi_status == "Uptrend":
                apply_uptrend_style(cell_status)
            elif rsi_status == "Downtrend":
                apply_downtrend_style(cell_status)
            else:
                apply_neutral_style(cell_status)
                
            # Chart link - initially set to "N/A" or "View Chart" based on availability
            if sheet_name in created_charts:
                cell_chart = results_ws.cell(row=data_row, column=6, value="View Chart")
                apply_cell_style(cell_chart)
            else:
                cell_chart = results_ws.cell(row=data_row, column=6, value="N/A")
                apply_cell_style(cell_chart)
            
            data_row += 1
            successful_sheets += 1
            logging.info(f"Successfully processed sheet: {sheet_name}")
            
        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {str(e)}")
            failed_sheets += 1
            continue
    
    # Add alternating row colors for readability
    header_row = current_row
    for row_idx in range(header_row + 1, data_row):
        if row_idx % 2 == 0:  # Even rows
            for col in range(1, 7):
                cell = results_ws.cell(row=row_idx, column=col)
                if not cell.fill or cell.fill.fill_type != "solid":
                    cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    
    # Add summary information
    summary_cell = results_ws.cell(row=current_row - 1, column=1, value=f"Total Stocks Analyzed: {data_row - header_row - 1}")
    summary_cell.font = Font(bold=True)
    results_ws.merge_cells(f'A{current_row-1}:F{current_row-1}')
    
    # Freeze panes for better navigation
    results_ws.freeze_panes = results_ws[f'A{header_row+1}']
    
    # Save the results workbook
    logging.info(f"Saving results workbook to {results_file}")
    results_wb.save(results_file)
    
    # Log how many charts were created
    logging.info(f"Created {len(created_charts)} charts out of {successful_sheets} successful sheets")
    
    # Add hyperlinks to charts and navigation buttons
    logging.info("Adding hyperlinks to charts and navigation buttons")
    add_chart_hyperlinks_and_navigation(results_file, output_dir, created_charts)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Analysis complete in {elapsed_time:.2f} seconds")
    logging.info(f"Successfully processed {successful_sheets} sheets, failed to process {failed_sheets} sheets")
    logging.info(f"Results saved to {results_file}")

def add_chart_hyperlinks_and_navigation(results_file, output_dir, created_charts=None):
    """Add hyperlinks to chart images and navigation buttons in the Excel file"""
    logging.info("Adding chart hyperlinks and navigation buttons to the Excel file")
    
    # Default to empty dict if created_charts is None
    if created_charts is None:
        created_charts = {}
    
    logging.info(f"Adding hyperlinks for {len(created_charts)} charts")
    
    if len(created_charts) == 0:
        logging.warning("No charts were created, skipping hyperlink creation")
        return
    
    wb = load_workbook(results_file)
    main_sheet = wb.active
    
    # Find the first data row (after headers)
    first_data_row = 0
    for row in range(1, main_sheet.max_row + 1):
        if main_sheet.cell(row=row, column=1).value == "Stock":
            first_data_row = row + 1
            break
            
    if first_data_row == 0:
        logging.error("Could not find data rows in the worksheet")
        return
    
    # Get the maximum row
    max_row = main_sheet.max_row
    logging.info(f"Processing {max_row-first_data_row+1} stock entries")
    
    # Count of charts processed
    charts_processed = 0
    
    # Create a sheet for each chart
    for row in range(first_data_row, max_row + 1):
        stock_name = main_sheet[f'A{row}'].value
        
        if not stock_name or stock_name not in created_charts:
            logging.debug(f"Skipping {stock_name} - no chart available")
            continue
            
        chart_path = created_charts[stock_name]
        
        if os.path.exists(chart_path):
            logging.info(f"Creating chart sheet for {stock_name}")
            # Create a new sheet for this stock's chart
            chart_sheet_name = f"{stock_name}_Chart"
            # Use truncated name if too long (Excel has a 31 character limit for sheet names)
            if len(chart_sheet_name) > 31:
                chart_sheet_name = chart_sheet_name[:28] + "..."
                
            chart_sheet = wb.create_sheet(title=chart_sheet_name)
            
            # Add a title to the chart sheet
            chart_sheet['A1'] = f"{stock_name} - RSI Analysis"
            chart_sheet['A1'].font = Font(bold=True, size=14)
            chart_sheet['A1'].alignment = Alignment(horizontal="center")
            
            # Merge cells for the title
            chart_sheet.merge_cells('A1:F1')
            
            # Add the chart image to the sheet
            logging.info(f"Adding chart image to sheet: {chart_sheet_name}")
            try:
                img = Image(chart_path)
                img.width = 800
                img.height = 500
                chart_sheet.add_image(img, 'A3')
                
                # Add "Go to Main Page" button at position O18 (right side of image at mid-height)
                button_cell = chart_sheet['O18'] = "Go to Main Page"
                apply_button_style(chart_sheet['O18'])
                chart_sheet['O18'].hyperlink = f"#'Analysis Results'!A1"
                
                # Make sure column O is wide enough for button text
                chart_sheet.column_dimensions['O'].width = 20
                
                # Create a hyperlink from the results sheet to the chart sheet
                logging.info(f"Creating hyperlink for {stock_name}")
                link_cell = main_sheet[f'F{row}']
                link_cell.value = "View Chart"
                link_cell.hyperlink = f"#{chart_sheet_name}!A1"
                apply_hyperlink_style(link_cell)
                
                charts_processed += 1
            except Exception as e:
                logging.error(f"Error adding chart image for {stock_name}: {str(e)}")
                # If there's an error, mark the link as not available
                main_sheet[f'F{row}'].value = "Chart Error"
        else:
            logging.warning(f"Chart not found for {stock_name}: {chart_path}")
    
    logging.info(f"Successfully processed {charts_processed} charts")
    logging.info(f"Saving workbook with hyperlinks and navigation to {results_file}")
    wb.save(results_file)
    logging.info("Chart hyperlinks and navigation buttons added successfully")

if __name__ == "__main__":
    logging.info("===== RSI ANALYSIS SCRIPT STARTED =====")
    
    # Find the latest data file
    latest_file = find_latest_data_file()
    
    if latest_file:
        logging.info(f"Processing file: {latest_file}")
        generate_analysis_report(latest_file)
    else:
        logging.error("No data files found matching the pattern 'shares_data_till_*.xlsx'")
    
    logging.info("===== RSI ANALYSIS SCRIPT COMPLETED =====")